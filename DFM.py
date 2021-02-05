import pandas as pd
import os
import openpyxl
import copy
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side, numbers
from openpyxl.drawing.image import Image
from openpyxl.utils import cell as ce
from openpyxl.utils.units import points_to_pixels as f2p
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU as p2e
from tqdm import tqdm

class DemographicFileMaker:

    def __init__(self, **args):
        ## initialize the object by specifying input and output files.
        self.image_src = args['image']

        self.raw_data_file = args['raw_data']
        self.raw_data_past_file = args['raw_data_past']
        self.item_code_file = args['item_code']
        self.demographics_file = args['demographics']
        self.demographics_past_file = args['demographics_past']
        self.heatmap_color_file = args['heatmap_color']
        self.benchmark_file = args['benchmark']
        self.leader_file = args['leader']
        self.output_source = args['output_folder']
        self.input_source = args['input_folder']
        
        self.performance_order = [
            "Exceptional",
            "Exceeded",
            "Achieved",
            "Improvement Needed",
            "On Leave",
            "No Rating",
        ]
    
    def setLeader(self, id, GM=False):
        self._leader_id = id
        self.GM = GM

    def readAllFiles(self):
        ## read files and save it in object data.
        self.origin_raw_data_pd = pd.read_excel(self.input_source + "/" + self.raw_data_file, engine="openpyxl")
        self.origin_raw_data_past_pd = pd.read_excel(self.input_source + "/" + self.raw_data_past_file, engine="openpyxl")
        self.item_code_pd = pd.read_excel(self.input_source + "/" + self.item_code_file, engine="openpyxl", sheet_name="ItemCodeSTAR")
        self.origin_category_pd =  pd.read_excel(self.input_source + "/" + self.item_code_file, engine="openpyxl", sheet_name="CurrentCategorySTAR")
        self.origin_demographics_pd = pd.read_excel(self.input_source + "/" + self.demographics_file, engine="openpyxl")
        self.demographics_past_pd = pd.read_excel(self.input_source + "/" + self.demographics_past_file, engine="openpyxl")
        self.heatmap_color_pd = pd.read_excel(self.input_source + "/" + self.heatmap_color_file, engine="openpyxl")
        self.benchmark_pd = pd.read_excel(self.input_source + "/" + self.benchmark_file, engine="openpyxl")
        self.leaders = pd.read_excel(self.input_source + "/" + self.leader_file, engine="openpyxl", sheet_name="Leader")
        self.GMs = pd.read_excel(self.input_source + "/" + self.leader_file, engine="openpyxl", sheet_name="GM")

    def calculateValues(self):
        
        ## do some process referred to individual leader ID.
        self._preProcess()
        self._prepareColumnsForID()

        item_list = []
        item_dict = {}
        category_list = []

        ## fill the above variables.
        for key in self.order_category:
            category_list.append(key)
            item_list.append([0, key])
            temp_list = []
            for item in self.category_pd.iloc[self._group_dict[key], 0]:
                item_list.append([1, item])
                temp_list.append(item)
            item_dict.update({key:temp_list})

        self._item_list = item_list

        ## do main calculation process by iterating over the category list.
        for category in tqdm(category_list, desc="iterating over the list of category"):
            sub_id_list = item_dict[category]

            ## filter the source data frame to have columns belonged to each category.
            self._filterResource(sub_id_list)

            ## based on the filtered source, calculate the values for each row.
            self._calculateEachRow(category)
       
    def writeOutput(self):

        ## specify the path of output file
        if self.GM:
            self.file_name = self.GM[:-4]
            self.output_path = "/" + self.file_name
        path = self.output_source + self.output_path + "/" + "2021-04 Employee Survey - " + self.file_name + " - Results by Demographics.xlsx"

        ## if the output file already exists, remove it.
        if os.path.exists(path):
            os.remove(path)

        ## make a folder to involve the output file.
        os.makedirs(self.output_source + self.output_path, exist_ok=True)

        ## and write the output file.
        self.book.save(path)

    def makeReport(self):

        ## eliminate columns of which respondents are lower than 4.
        criteria_dict = copy.deepcopy(self.precious_dict)
        for key in criteria_dict:
            if key != "":
                for sub_key in criteria_dict[key]:
                    try:
                        if self.first_row[key + sub_key] < 4:
                            del self.precious_dict[key][sub_key]
                    except:
                        pass

        ## eliminate section of which segments are lower than 2.
        criteria_dict = copy.deepcopy(self.precious_dict)
        for key in criteria_dict:
            if len(criteria_dict[key]) < 2 and key != "":
                del self.precious_dict[key]

        ## make a content to be displayed in the picture position.
        field_picture_postion = str(round(self._participated / self._invited * 100)) + "% Participation Rate" + \
            "\n" + str(self._participated) + ' / ' + str(self._invited) + "\n" + "(Participated / Invited)"

        ## prepare image.
        img  = Image(self.input_source + "/" + self.image_src)
        img.height = 90
        img.width = 110

        ## calculate total rows that will be placed in our output file.
        total_rows = 4 + len(self._item_list)

        ## make a workbook and sheet.
        self.book = openpyxl.Workbook()
        sheet = self.book.active

        ## set styles like font, color, direction, border...
        ft = Font(name="Arial", size=8)
        ft_bold = Font(name="Arial", size=8, bold=True)
        light_na_font = Font(name="Arial", size=8, color="999999")
        bold_na_font = Font(name="Arial", size=8, color="999999", bold=True)
        vertical = Alignment(textRotation=90, horizontal='center')
        center_alignment = Alignment(horizontal='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        grey_back = PatternFill("solid", fgColor="EEEEEE")
        white_back = PatternFill("solid", fgColor="FFFFFF")
        side = Side(style='thin', color="CCCCCC")
        thin_border = Border(left=side,
                     right=side,
                     top=side,
                     bottom=side)

        ## merge all needed columns and rows.
        col_num = 1
        skip_list = []
        for key in self.precious_dict:
            if len(self.precious_dict[key]) == 0:
                continue
            delta = len(self.precious_dict[key]) - 1
            if col_num == 1:
                delta += 1
            end_col = delta + col_num
            cell = sheet.cell(row=1, column=col_num)
            cell.font = ft
            cell.value = "" if key == "delta" else key
            sheet.merge_cells(start_row=1, start_column=col_num, end_row=2, end_column=end_col)
            col_num += delta + 2

            ## make a skip_list which will be used to make a empty column between groups.
            skip_list.append(col_num - 1)

        ## set empty cells white.
        for row in range(1, total_rows + 1 + 1):
            for col in range(1, col_num - 2 + 1 + 1):
                sheet.cell(row=row, column=col).fill = white_back

        ## prepare the whole data to be placed in the sheet.
        frames = []
        for key in tqdm(self.precious_dict, desc="prepare the whole data"):
            sub_dict = self.precious_dict[key]
            
            ## prepare and write the first column data.
            if key == '':
                _list = []
                _list.append([field_picture_postion, 1])
                _list.append(["Number of Respondents (incl. N/A)", 1])
                for criteria, item in self._item_list:
                    if criteria == 0:
                        _list.append([item, 0])
                    elif criteria == 1:
                        _list.append([self._item_pd[self._item_pd["Item ID"] == item]["Short Text [2020 onward]"].values[0], 1])
                
                ## write the first column data and set styles.
                for index, item in enumerate(_list):
                    cell = sheet.cell(row=3 + index, column=1)
                    if item[1] == 0:
                        cell.font = ft_bold
                    else:
                        cell.font = ft

                    cell.border = thin_border
                    cell.fill = grey_back
                    cell.alignment = right_alignment
                    cell.alignment = cell.alignment.copy(horizontal="left")

                    ## set value to cell.
                    cell.value = item[0]

                ## autofit the first column's width.
                _ = 0
                for item in _list[1:]:
                    if len(item[0]) > _:
                        _ = len(item[0])

                sheet.column_dimensions["A"].width = _ / 1.6

            ## prepare the rest of data to fill other columns
            for sub_key in sub_dict:

                len_sub_key = 0
                if len(sub_key) > len_sub_key:
                    len_sub_key = len(sub_key)

                sub_item = sub_dict[sub_key]
                _list = []
                if sub_key == "Δ Your Org (2018)" and self.logic == 1:
                    _list.append(["Δ Gilead Overall (2018)", 2])
                elif sub_key == "Δ Your Org (2018)" and self.GM:
                    _list.append(["Δ " + self.GM + " (2018)", 2])
                elif sub_key == "Parent Group":
                    _list.append([self._supervisor_last_name + " Org", 2])
                elif sub_key == "Your Org (2020)" and self.GM:
                    _list.append([self.GM + " (2020)", 2])
                elif sub_key.endswith("Office"):
                    _list.append(["Office", 2])
                else:
                    _list.append([sub_key, 2])
                _list.append([self.first_row[key + sub_key], 1])

                ## prepare history column.
                if sub_key == "Δ Your Org (2018)":
                    for criteria, item in self._item_list:
                        origin_item = item
                        if criteria == 1:
                            origin_item = self._item_pd[self._item_pd["Item ID"] == item]["Unique Item Code"].values[0]
                            item = self._get_past_field_name_by_current_name(origin_item)
                        try:
                            _ = sub_item[item]
                            if _[0] != "N/A":
                                _[0] = self.precious_dict[""]["Your Org (2020)" if self.logic != 1 else "Gilead Overall"][origin_item][0] - _[0]
                            _list.append(_)
                        except:
                            _list.append(["N/A", 1 if criteria else 0])
                ## prepare benchmark column.
                elif sub_key == "Δ External":
                    for criteria, item in self._item_list:
                        origin_item = item
                        if criteria == 1:
                            origin_item = self._item_pd[self._item_pd["Item ID"] == item]["Unique Item Code"].values[0]
                            item = self._get_benchmark_field_name_by_current_name(origin_item)
                        try:
                            _ = sub_item[item]
                            if _[0] != "N/A":
                                _[0] = self.precious_dict[""]["Your Org (2020)" if self.logic != 1 else "Gilead Overall"][origin_item][0] - _[0]
                            _list.append(_)
                        except:
                            _list.append(["N/A", 1 if criteria else 0])
                else:
                    for criteria, item in self._item_list:
                        if criteria == 1:
                            item = self._item_pd[self._item_pd["Item ID"] == item]["Unique Item Code"].values[0]

                        _list.append(sub_item[item])
                    
                frames.append(_list)
        
        ## this method is used to make a empty column when write columns.
        def get_column_number(number):

            for num in skip_list:
                if number >= num:
                    number += 1
            return number
        
        ## below method is used to calculate delta.
        def get_delta(first, second):
            return round(second * 100) - round(first * 100)

        gilead_org = frames[0]

        ## write the rest of the columns and set styles.
        for col_index, column in enumerate(tqdm(frames, desc="formating and styling")):
            for row_index, item in enumerate(column):
                cell = sheet.cell(row=3 + row_index, column = get_column_number(2 + col_index))

                ## set borders of all cells.
                cell.border = thin_border

                ## set background grey of the row->4.
                if row_index == 1:
                    cell.fill = grey_back
                    cell.number_format = numbers.BUILTIN_FORMATS[3]

                ## if the cell is placed in category row, set bold font style. Otherwise set general font style.
                if item[1] == 0:
                    cell.font = ft_bold
                else:
                    cell.font = ft
                cell.alignment = right_alignment

                ## set corresponding styles to row->3
                if item[1] == 2:
                    cell.alignment = center_alignment
                    cell.alignment = vertical
                    cell.fill = grey_back
                
                ## set value to cell.
                cell.value = item[0]

                ## make "N/A" cell lightgrey.
                if row_index >= 1:
                    if item[0] == "N/A":
                        cell.alignment = right_alignment
                        cell.font = bold_na_font if item[1] == 0 else light_na_font

                ## set background color and set format of percentage to rows below 5.
                if row_index >= 2:

                    ## set format percentage.
                    cell.number_format = numbers.FORMAT_PERCENTAGE

                    ## compare the rest of the columns with your org and set background.
                    if col_index >= self.logic:
                        try:
                            # cell.fill = PatternFill("solid", fgColor=self._get_color(get_delta(your_org[row_index][0], item[0])))
                            cell.fill = PatternFill("solid", fgColor=self._get_color(get_delta(frames[self.logic - 1][row_index][0], item[0])))
                        except:
                            ## this skip the case of N/A
                            pass

                    ## compare your org (2020) with parent org and set background.
                    elif col_index == self.logic - 1 and self.logic >= 2:
                        try:
                            cell.fill = PatternFill("solid", fgColor=self._get_color(get_delta(frames[self.logic - 2][row_index][0], item[0])))
                        except:
                            ## this skip the case of N/A
                            pass

                    ## compare parent org with gilead org and set background.
                    elif col_index == 1 and self.logic == 3:
                        try:
                            cell.fill = PatternFill("solid", fgColor=self._get_color(get_delta(gilead_org[row_index][0], item[0])))
                        except:
                            ## this skip the case of N/A
                            pass
                    
                    ## set background color about history data.
                    try:
                        _ = len(self.precious_dict["Direct Reports (as of April 24, 2018)"].keys())
                    except:
                        _ = 0
                    if col_index == self.logic + _ or col_index == self.logic + 1 + _:
                        try:
                            cell.fill = PatternFill("solid", fgColor=self._get_color(round(item[0] * 100)))
                        except:
                            ## this skip the case of N/A
                            pass
                    
                    ## force adding PLUS symbol.
                    if column[0][0][0] == "Δ":
                        try:
                            if item[0] > 0:
                                cell.number_format = "+0%"
                        except:
                            pass

        ## set columns' width and empty columns' width.
        for i in range(2, col_num - 2 + 1):
            name = ce.get_column_letter(i)
            if i in skip_list:
                sheet.column_dimensions[name].width = 0.8
            else:
                sheet.column_dimensions[name].width = 4.5
        
        ## set all rows' height.
        for i in range(1, total_rows + 1):
            if i == 3:
                sheet.row_dimensions[i].height = 95
                if len_sub_key > 20:
                    sheet.row_dimensions[i].height = len_sub_key / 20 * 95
            else:
                sheet.row_dimensions[i].height = 10.2

        ## freeze panes.
        sheet.freeze_panes = ce.get_column_letter(self.logic + 2) + '5'

        ## push content right in A3, A4.
        sheet['A3'].alignment = right_alignment
        sheet['A4'].alignment = right_alignment

        ## insert picture and set background white
        sheet['A3'].fill = white_back
        size = XDRPositiveSize2D(p2e(115), p2e(95))
        marker = AnchorMarker(col=0, colOff=p2e(0), row=2, rowOff=p2e(10))
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        
        sheet['A3'].alignment = sheet['A3'].alignment.copy(wrapText=True, vertical="bottom")
        sheet.add_image(img)
    
    def _get_names_from_field(self, field_list):

        keys = [field[0] for field in field_list]
        return keys


    def _filterResource(self, id_list):

        filter_item = self._item_pd[self._item_pd["Item ID"].isin(id_list)]["Unique Item Code"].tolist()

        filter_item.insert(0, 'ExternalReference')
        self._filtered_raw_data = self.raw_data_pd[filter_item]

        ## about history data.
        self._helper_past = False
        filter_past_item = []
        for index, item in enumerate(filter_item):
            if index == 0:
                filter_past_item.append(item)
            else:
                new_item = self._get_past_field_name_by_current_name(item)
                if new_item != None:
                    filter_past_item.append(new_item)
                else:
                    self._helper_past = True
        
        self._filtered_raw_past_data = self.raw_data_past_pd[filter_past_item]

        ## about benchmark data.
        self._helper_benchmark = False
        filter_benchmark_item = []
        for index, item in enumerate(filter_item):
            if index == 0:
                filter_benchmark_item.append(item)
            else:
                new_item = self._get_benchmark_field_name_by_current_name(item)
                if new_item:
                    filter_benchmark_item.append(new_item)
                else:
                    self._helper_benchmark = True
        
        self._filtered_benchmark_value = self.benchmark_pd[self.benchmark_pd["Unique Item Code"].isin(filter_benchmark_item)]

    def _get_color(self, delta):
        if abs(delta) > 25:
            delta = delta // abs(delta) * 25
        series = self.heatmap_color_pd[self.heatmap_color_pd["Delta"] == delta]
        return str(hex(series["R"].values[0]))[2:] + str(hex(series["G"].values[0]))[2:] + str(hex(series["B"].values[0]))[2:]

    def _preProcess(self):

        ## make a item group and filter the able source.
        self._item_pd = self.item_code_pd[self.item_code_pd["Type ID"] == "T01"]

        self._benchmark_item_pd = self._item_pd[self._item_pd["External Benchmark"] == "e"]
        self._item_pd = self._item_pd[self._item_pd["External Benchmark"] == "i"]

        self._rest_item_pd = self._item_pd[~self._item_pd["Unique Item Code"].isin(self.origin_raw_data_pd.columns.values)].reset_index(drop=True)
        self._item_pd = self._item_pd[self._item_pd["Unique Item Code"].isin(self.origin_raw_data_pd.columns.values)].reset_index(drop=True)
        self._rest_item_pd = self._rest_item_pd[self._rest_item_pd["Unique Item Code"].isin(self.origin_raw_data_past_pd.columns.values)].reset_index(drop=True)
        self._both_item_pd = self._item_pd[self._item_pd["Unique Item Code"].isin(self.origin_raw_data_past_pd.columns.values)].reset_index(drop=True)

        self.category_pd = self.origin_category_pd[self.origin_category_pd["Item ID in 2020 Survey"].isin(self._item_pd["Item ID"].tolist())]
        self.category_pd = self.category_pd.drop_duplicates(subset=["Item ID in 2020 Survey"]).reset_index(drop=True)
        self.order_category = self.category_pd.drop_duplicates(subset=["2020 Category"]).loc[:, "2020 Category"].tolist()

        self._group_dict = self.category_pd.groupby(["2020 Category"]).groups

        self.raw_data_pd = self.origin_raw_data_pd.iloc[2:].reset_index(drop=True)

        ## filter the able source from history file.
        self.raw_data_past_pd = self.origin_raw_data_past_pd.iloc[2:].reset_index(drop=True)

        ## Convert numeric values into favorable or not.
        ## [1, 2, 3] -> 0, [4, 5] -> 1, [6, -99, ''] -> ''
        for field in self._item_pd["Unique Item Code"]:
            new_list = []
            for item in self.raw_data_pd[field].tolist():
                if item < 4 and item > 0:
                    new_list.append(0)
                elif item >= 4 and item <= 5:
                    new_list.append(1)
                else:
                    new_list.append('')
            self.raw_data_pd[field] = new_list
            
            ## do the same process about history data.
            try:
                field = self._get_past_field_name_by_current_name(field)
                new_list_past = []
                for item in self.raw_data_past_pd[field].tolist():
                    if item < 4 and item > 0:
                        new_list_past.append(0)
                    elif item >= 4 and item <= 5:
                        new_list_past.append(1)
                    else:
                        new_list_past.append('')
                self.raw_data_past_pd[field] = new_list_past
            except:
                pass

        ## new feature -> process A/B pair.
        pairs = self._item_pd[self._item_pd["AB Code"].isin(["A", "B"])].reset_index(drop=True)
        pairs_dict = pairs.groupby(["Item ID"]).groups

        for key in pairs_dict:

            _list = pairs_dict[key]
            pairs_row = pairs.iloc[_list, :]
            item_list = pairs_row["Unique Item Code"].tolist()
            text_list = pairs_row["Short Text [2020 onward]"].tolist()
            new_text = "/".join(text_list)
            self._item_pd = self._item_pd[~(self._item_pd["Unique Item Code"] == item_list[1])].reset_index(drop=True)
            _index = self._item_pd[self._item_pd["Unique Item Code"] == item_list[0]].index
            self._item_pd.loc[_index, "Short Text [2020 onward]"] = new_text

            column_pair = self.raw_data_pd[item_list]
            _list = []
            for index in range(len(column_pair.index)):
                _row = column_pair.iloc[index, :].tolist()
                for val in _row:
                    if type(val) == type(0):
                        _list.append(val)
                        break
                else:
                    _list.append('')

            _pd = pd.DataFrame(_list, columns=[item_list[0]])
            self.raw_data_pd = self.raw_data_pd.drop(columns=[item_list[1]])
            self.raw_data_pd[item_list[0]] = _pd

        self.demographics_pd = self.origin_demographics_pd

        ## convert the demographics data to include only answered entries.
        self._answered_demographics_data = self.demographics_pd[self.demographics_pd.loc[:, "MCJ ID"].isin(self.raw_data_pd['ExternalReference'].tolist())].reset_index(drop=True)
        self._gilead_org = self._answered_demographics_data

        ## do the same process about history data.
        self._answered_demographics_past_data = self.demographics_past_pd[self.demographics_past_pd.loc[:, "MCJ ID"].isin(self.raw_data_past_pd['ExternalReference'].tolist())].reset_index(drop=True)
    
        if self.GM:
            self._gm_demographics_data = self._answered_demographics_data[self._answered_demographics_data[self.GM] == 1].reset_index(drop=True)
            self._gm_demographics_past_data = self._answered_demographics_past_data[self._answered_demographics_past_data[self.GM] == 1].reset_index(drop=True)
        
    def _get_past_field_name_by_current_name(self, field_name):
        _item_id = self._item_pd[self._item_pd["Unique Item Code"] == field_name]["Item ID"].values[0]
        try:
            new_field_name = self._rest_item_pd[self._rest_item_pd["Item ID"] == _item_id]["Unique Item Code"].values[0]
        except:
            new_field_name = None

        if new_field_name == None:
            try:
                new_field_name = self._both_item_pd[self._both_item_pd["Unique Item Code"] == field_name]["Unique Item Code"].values[0]
            except:
                new_field_name = None

        return new_field_name

    def _get_benchmark_field_name_by_current_name(self, field_name):
        _item_id = self._item_pd[self._item_pd["Unique Item Code"] == field_name]["Item ID"].values[0]
        try:
            field_name = self._benchmark_item_pd[self._benchmark_item_pd["Item ID"] == _item_id]["Unique Item Code"].values[0]
        except:
            field_name = None
        
        return field_name

    def _prepareColumnsForID(self):
        ## find the supervisor level of the given leader id.

        self.first_row = {}
        _temp = "Supervisor Level {} MCJ ID"

        leader_entry = self.origin_demographics_pd[self.origin_demographics_pd.loc[:, "MCJ ID"] == self._leader_id]

        if self._leader_id == 999999:
            self.logic = 1
            leader_level = 1
        else:
            for index, level in enumerate(leader_entry.loc[:, _temp.format(2) : _temp.format(10)]):
                if (leader_entry[level] == self._leader_id).tolist()[0]:
                    leader_level = index + 2
                    break
        
        self.output_path = "/" + leader_entry["MCJ Name"].values[0]
        self.file_name = leader_entry["MCJ Last Name"].values[0]

        supervisor_level = leader_level - 1
        direct_level = leader_level + 1

        if leader_level >= 3:
            self.logic = 3
            self._supervisor_id = leader_entry.loc[:, _temp.format(supervisor_level)].values[0]
            _supervisor_entry = self.origin_demographics_pd[self.origin_demographics_pd.loc[:, "MCJ ID"] == self._supervisor_id]
            self._supervisor_last_name = _supervisor_entry["MCJ Last Name"].values[0]

            ## get Parent group.
            self._parent_org = self._answered_demographics_data[self._answered_demographics_data.loc[:, _temp.format(supervisor_level)] == self._supervisor_id].reset_index(drop=True)

        elif leader_level == 2:
            self.logic = 2

        ## get Your Org data
        if self.logic == 1:
            self._your_org = self._answered_demographics_data
            self._invited = len(self.demographics_pd.index)
        else:
            if self.GM:
                self._your_org = self._gm_demographics_data
                self._invited = len(self.origin_demographics_pd[self.origin_demographics_pd[self.GM] == 1].index)
            else:
                self._your_org = self._answered_demographics_data[self._answered_demographics_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True)
                self._invited = len(self.demographics_pd[self.demographics_pd.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True).iloc[:, 0])
        
        ## calculate nums of participated
        self._participated = len(self._your_org.loc[:, "MCJ ID"])
        
        ## get your history group
        if self.GM:
            self._your_past_org = self._gm_demographics_past_data
        else:
            self._your_past_org = self._answered_demographics_past_data[self._answered_demographics_past_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True)


        ## make direct report fields.
        self._direct_report_field = []
        temp_org = self._your_org[~(self._your_org.loc[:, "MCJ ID"] == self._leader_id)].reset_index(drop=True)

        if self.GM:
            temp_org = temp_org[temp_org.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True)

        _dict = temp_org.groupby(_temp.format(direct_level)).groups
        for key in _dict:
            try:
                self._direct_report_field.append([self.origin_demographics_pd[self.origin_demographics_pd.loc[:, "MCJ ID"] == key].loc[:, "MCJ Name"].values[0], _dict[key]])
            except:
                pass

        ## make grade group fields.
        self._grade_group_fields = []
        _dict = self._your_org.groupby("Pay Grade Group").groups
        for key in _dict:
            self._grade_group_fields.append([key, _dict[key]])

        ## make tenure group fields.
        self._tenure_group_fields = []
        _dict = self._your_org.groupby("Length of Service Group").groups
        for key in _dict:
            if not key == "15+ Years":
                self._tenure_group_fields.append([key, _dict[key]])
        try:
            self._tenure_group_fields.append(["15+ Years", _dict["15+ Years"]])
        except:
            pass
        
        ## make performance rating fields.
        self._performance_rating_fields = []
        _dict = self._your_org.groupby("2019 Performance Rating").groups
        _list = _dict.keys()
        for key in self.performance_order:
            try:
                self._performance_rating_fields.append([key, _dict[key]])
                _list.remove(key)
            except:
                pass
        for key in _list:
            self._performance_rating_fields.append([key, _dict[key]])

        ## make talent cordinate fields.
        self._talent_cordinate_fields = []
        _dict = self._your_org.groupby("2020 Talent Coordinate").groups
        for key in _dict:
            self._talent_cordinate_fields.append([key, _dict[key]])
        
        ## make gender fields.
        self._gender_fields = []
        _dict = self._your_org.groupby("Gender").groups
        for key in _dict:
            self._gender_fields.append([key, _dict[key]])

        ## make Ethnicity fields.
        self._ethnicity_fields = []
        _dict = self._your_org.groupby("Ethnicity (US)").groups
        for key in _dict:
            self._ethnicity_fields.append([key, _dict[key]])

        ## make age group fields.
        self._age_fields = []
        _dict = self._your_org.groupby("Age Group").groups
        for key in _dict:
            self._age_fields.append([key, _dict[key]])

        ## make country fields.
        self._country_fields = []
        _dict = self._your_org.groupby("Country").groups
        for key in _dict:
            self._country_fields.append([key, _dict[key]])

        ## make kite fields.
        self._kite_fields = []
        _dict = self._your_org.groupby("Kite Employee Flag").groups
        for key in _dict:
            self._kite_fields.append([key, _dict[key]])

        ## make office type fields.
        self._office_fields = []
        _dict = self._your_org.groupby("Office Type").groups
        for key in _dict:
            self._office_fields.append([key, _dict[key]])

        self.index_match = {
            "": ["Gilead Overall", "Parent Group", "Your Org (2020)"],
            "Direct Reports (as of April 24, 2018)": sorted(self._get_names_from_field(self._direct_report_field)),
            "delta": ["Δ Your Org (2018)", "Δ External"],
            "Grade Group": self._get_names_from_field(self._grade_group_fields),
            "Tenure Group": self._get_names_from_field(self._tenure_group_fields),
            "Office Type": self._get_names_from_field(self._office_fields),
            "2019 Performance Rating": self._get_names_from_field(self._performance_rating_fields),
            "2020 Talent Coordinate": self._get_names_from_field(self._talent_cordinate_fields),
            "Gender": self._get_names_from_field(self._gender_fields),
            "Ethnicity (US)": self._get_names_from_field(self._ethnicity_fields),
            "Age Group": self._get_names_from_field(self._age_fields),
            "Country": self._get_names_from_field(self._country_fields),
            "Kite": self._get_names_from_field(self._kite_fields),
        }

        if self.logic == 2:
            self.index_match.update({"": ["Gilead Overall", "Your Org (2020)"]})
        elif self.logic == 1:
            self.index_match.update({"": ["Gilead Overall"]})

        ## make and initiate a dict to save all calculated data.
        self.precious_dict = {}
        for first_index in self.index_match:
            _ = {}
            for item in self.index_match[first_index]:
                _.update({item: {}})
            self.precious_dict.update({first_index: _})

    def _calculateEachRow(self, item):
        
        ## calculate Gilead overall %s
        _dict, lens = self._calculateOverall(self._gilead_org, item)
        self.precious_dict[""]["Gilead Overall"].update(_dict)
        self.first_row.update({"Gilead Overall": lens})

        ## calculate Parent Group %s
        if self.logic == 3:
            _dict, lens = self._calculateOverall(self._parent_org, item)
            self.precious_dict[""]["Parent Group"].update(_dict)
            self.first_row.update({"Parent Group": lens})

        ## calculate Your Org(2018) %s
        if self.logic >= 2:
            _dict, lens = self._calculateOverall(self._your_org, item)
            self.precious_dict[""]["Your Org (2020)"].update(_dict)
            self.first_row.update({"Your Org (2020)": lens})

        ## calculate Δ Your Org (2018) %s
        if len(self._filtered_raw_past_data.columns) > 1:
            _dict, lens = self._calculateOverall(self._your_past_org, item, history=True)
            self.precious_dict["delta"]["Δ Your Org (2018)"].update(_dict)
            self.first_row.update({"deltaΔ Your Org (2018)": lens})

        ## calculate Δ External %s
        if len(self._filtered_benchmark_value.columns) > 1:
            self.first_row.update({"deltaΔ External": "N/A"})
            _dict = {}
            for index, row in self._filtered_benchmark_value.iterrows():
                value = row["External - CAmp Biotechnology & Medical Devices 2019"]
                if not value:
                    _dict.update({row["Unique Item Code"]: [value, 1]})
                else:
                    _dict.update({row["Unique Item Code"]: [round(value, 2), 1]})
            _dict.update({item: ["N/A", 0]})
            self.precious_dict["delta"]["Δ External"].update(_dict)

        ## calculate Direct reports %s
        self._calculateSubFields(self._direct_report_field, "Direct Reports (as of April 24, 2018)", item)
        
        ## calcualte Grade Group %s
        self._calculateSubFields(self._grade_group_fields, "Grade Group", item)

        ## calcualte Tenure Group %s
        self._calculateSubFields(self._tenure_group_fields, "Tenure Group", item)

        ## calculate Performance Rating %s
        self._calculateSubFields(self._performance_rating_fields, "2019 Performance Rating", item)

        ## calculate Talent Coordinate %s
        self._calculateSubFields(self._talent_cordinate_fields, "2020 Talent Coordinate", item)

        ## calculate Gender %s
        self._calculateSubFields(self._gender_fields, "Gender", item)

        ## calculate Ethnicity (US) %s
        self._calculateSubFields(self._ethnicity_fields, "Ethnicity (US)", item)

        ## calculate Age Group %s
        self._calculateSubFields(self._age_fields, "Age Group", item)

        ## calculate Country %s
        self._calculateSubFields(self._country_fields, "Country", item)

        ## calculate Kite %s
        self._calculateSubFields(self._kite_fields, "Kite", item)

        ## calculate office %s
        self._calculateSubFields(self._office_fields, "Office Type", item)

    def _calculateOverall(self, dataframe, item, history=False):

        ## calcualte overall fields.
        ids = dataframe.iloc[:, 0]
        # nums = len(ids)
        if history:
            working_pd = self._filtered_raw_past_data[self._filtered_raw_past_data["ExternalReference"].isin(ids)].reset_index(drop=True)
        else:
            working_pd = self._filtered_raw_data[self._filtered_raw_data["ExternalReference"].isin(ids)].reset_index(drop=True)
        nums = len(working_pd.iloc[:, 0])
        return self._get_sum(working_pd.iloc[:, 1:], nums, item, history), len(ids)
    
    def _calculateSubFields(self, dataframe, column_name, item):

        ## calculate the sum of all sub fields except overall fields.
        for field in dataframe:
            dictionary = self.precious_dict[column_name]
            _list = field[1]
            ids = self._your_org.iloc[_list, 0].tolist()
            nums = len(ids)
            working_pd = self._filtered_raw_data[self._filtered_raw_data["ExternalReference"].isin(ids)].reset_index(drop=True)
            _dict = self._get_sum(working_pd.iloc[:, 1:], nums, item)
            dictionary[field[0]].update(_dict)
            self.first_row.update({column_name + field[0]: len(ids)})
    
    def _get_sum(self, data, nums, item, implicity=False):

        ## calculate the sum of favorable scores.
        _dict = {}
        determine_parent_na = False

        for ind in range(len(data.columns)):
            _ = data.iloc[:, ind]
            is_empty_column = True
            lens = nums
            sub = 0
            count_valid = 0
            for __ in _:
                if __ != '':
                    is_empty_column = False
                    sub += __
                    count_valid += 1
                else:
                    lens -= 1
            if is_empty_column:
                _dict.update({data.columns.values[ind]: ["N/A", 1]})
                determine_parent_na = True
            else:
                if count_valid >= 4:
                    _dict.update({data.columns.values[ind]: [round(sub / lens, 2), 1]})
                else:
                    _dict.update({data.columns.values[ind]: ["N/A", 1]})
                    determine_parent_na = True
        
        if determine_parent_na:
            _dict.update({item: ["N/A", 0]})
        else:
            total_lens = nums
            total = 0
            cols = len(data.columns)
            for ind in range(nums):
                _series = data.iloc[ind, :]
                sub_sum = 0
                _is_nan = False
                for val in _series:
                    if val != '':
                        sub_sum += val
                    else:
                        _is_nan = True
                        total_lens -= 1
                        break

                if not _is_nan:
                    total += sub_sum / cols

            if total_lens < 4:
                _dict.update({item: ["N/A", 0]})
            else:
                _dict.update({item: [round(total / total_lens, 2), 0]})

        if implicity:
                _dict.update({item: ["N/A", 0]})
        return _dict

if __name__ == "__main__":

    ## Create a object.
    init_data = {
        'leader': "List of Leaders and GMs 2021-01-26.xlsx",
        'raw_data': "Qualtrics Survey Export Sample New.xlsx",
        'item_code': "Item Code SHARE 2021-01-23.xlsx",
        'demographics': "Demographics File Sample 2021-01-26.xlsx",
        'heatmap_color': "Heatmap Colors.xlsx",
        'raw_data_past': "2018 Employee Survey Responses Sample 2021-01-23.xlsx",
        'demographics_past': "2018 Demographics File Sample 2021-01-26.xlsx",
        'benchmark': "External Benchmarks.xlsx",
        'output_folder': "./output",
        'input_folder': "./input",
        'image': "/image.png",
    }

    dfm = DemographicFileMaker(**init_data)

    ## read all needed files.
    dfm.readAllFiles()

    total_ids = len(dfm.leaders.index) + len(dfm.GMs.index)

    for index in tqdm(range(total_ids), desc="total process"):
        if index < len(dfm.leaders.index):
            dfm.setLeader(dfm.leaders.iloc[index, :].values[0])
        else:
            row = dfm.GMs.iloc[index - len(dfm.leaders.index), :]
            dfm.setLeader(row["GM ID"], row["GM Org"])
    
        ## do main process to calculate report.
        dfm.calculateValues()

        ## make report dataframe to output.
        dfm.makeReport()

        ## write output file (Mockup STAR.xlsx)
        dfm.writeOutput()
    print("complete!")