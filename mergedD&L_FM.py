import pandas as pd
import os
import openpyxl
import copy
from openpyxl.styles import Font, Alignment, PatternFill, Color, Border, Side, numbers
from openpyxl.drawing.image import Image
from openpyxl.utils import cell as ce
from openpyxl.utils.units import points_to_pixels as f2p
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU as p2e
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
from decimal import Decimal
from tqdm import tqdm

class DemographicFileMaker:

    def __init__(self, **args):
        ## initialize the object by specifying input and output files.
        self.image_src = args['image']

        self.raw_data_file = args['raw_data']
        self.raw_data_past_file = args['raw_data_past']
        self.item_code_file = args['item_code']
        self.demographics_file = args['demographics']
        self.demographics_past_file = args['demographics_past']
        self.heatmap_color_file = args['heatmap_color']
        self.benchmark_file = args['benchmark']
        self.leader_file = args['leader']
        self.output_source = args['output_folder']
        self.input_source = args['input_folder']
        self.how2use_file = args['how to use']
        self.GM_levels_file = args['gm_levels']
        
        self.performance_order = [
            "Exceptional",
            "Exceeded",
            "Achieved",
            "Improvement Needed",
            "On Leave",
            "No Rating",
        ]

        self.current_year = self.demographics_file[:4]
        self.past_year = self.demographics_past_file[:4]

        self.GM_region_human_parentorg = 0
    
    def setLeader(self, id, GM=False, site_lead=False):
        self.use_affiliate = False
        self.affiliate_second = False

        self._leader_id = id
        self.GM = GM
        self.site_lead = site_lead
            
    def setGMParentFlag(self, value):
        self.GM_region_human_parentorg = value

    def readAllFiles(self):
        ## read files and save it in object data.
        self.origin_raw_data_pd = pd.read_excel(self.input_source + "/" + self.raw_data_file, engine="openpyxl")
        self.origin_raw_data_past_pd = pd.read_excel(self.input_source + "/" + self.raw_data_past_file, engine="openpyxl")
        self.item_code_pd = pd.read_excel(self.input_source + "/" + self.item_code_file, engine="openpyxl", sheet_name="ItemCodeSTAR")
        self.origin_category_pd =  pd.read_excel(self.input_source + "/" + self.item_code_file, engine="openpyxl", sheet_name="CurrentCategorySTAR")
        self.origin_demographics_pd = pd.read_excel(self.input_source + "/" + self.demographics_file, engine="openpyxl")
        self.demographics_past_pd = pd.read_excel(self.input_source + "/" + self.demographics_past_file, engine="openpyxl")
        self.heatmap_color_pd = pd.read_excel(self.input_source + "/" + self.heatmap_color_file, engine="openpyxl")
        self.benchmark_pd = pd.read_excel(self.input_source + "/" + self.benchmark_file, engine="openpyxl")
        self.leaders = pd.read_excel(self.input_source + "/" + self.leader_file, engine="openpyxl", sheet_name="Leader")
        self.GMs = pd.read_excel(self.input_source + "/" + self.leader_file, engine="openpyxl", sheet_name="GM")
        self.site_leads = pd.read_excel(self.input_source + "/" + self.leader_file, engine="openpyxl", sheet_name="Site Leader")
        self.GM_levels = pd.read_excel(self.input_source + "/" + self.GM_levels_file, engine="openpyxl")

    def calculateValues(self):
        
        ## do some process referred to individual leader ID.
        self._preProcess()
        _ = self._prepareColumnsForID()
        if _:
            return _

        item_list = []
        item_dict = {}
        category_list = []

        ## fill the above variables.
        for key in self.order_category:
            category_list.append(key)
            item_list.append([0, key])
            temp_list = []
            for item in self.category_pd.iloc[self._group_dict[key], 0]:
                item_list.append([1, item])
                temp_list.append(item)
            item_dict.update({key:temp_list})

        self._item_list = item_list

        ## do main calculation process by iterating over the category list.
        for category in tqdm(category_list, desc="iterating over the list of category"):
            sub_id_list = item_dict[category]

            ## filter the source data frame to have columns belonged to each category.
            self._filterResource(sub_id_list)

            ## based on the filtered source, calculate the values for each row.
            self._calculateEachRow(category)
       
    def writeOutput(self):

        ## specify the path of output file
        if self.GM:
            self.file_name = self.GM[:-4]
            self.output_path = "/" + self.file_name
        if self.site_lead:
            self.file_name = self.site_lead.replace(" / ", " ")
            self.output_path = "/" + self.file_name
        path = self.output_source + self.output_path + "/" + "2021-04 Global Employee Survey - " + self.file_name + " - Demographic Trends.xlsx"


        ## if the output file already exists, remove it.
        if os.path.exists(path):
            os.remove(path)

        ## make a folder to involve the output file.
        os.makedirs(self.output_source + self.output_path, exist_ok=True)

        ## and write the output file.
        self.book.save(path)

    def getWorkBook(self):
        return self.book

    def makeReport(self):

        ## eliminate columns of which respondents are lower than 4.
        criteria_dict = copy.deepcopy(self.precious_dict)
        for key in criteria_dict:
            if key != "":
                for sub_key in criteria_dict[key]:
                    try:
                        if self.first_row[key + sub_key] < 4:
                            del self.precious_dict[key][sub_key]
                    except:
                        pass

        ## eliminate section of which segments are lower than 2.
        criteria_dict = copy.deepcopy(self.precious_dict)
        for key in criteria_dict:
            if len(criteria_dict[key]) < 2 and key != "":
                del self.precious_dict[key]

        def get_percent(numerator, denominator):
            try:
                value = str(round(numerator / denominator * 100)) + "% Participation Rate"
            except:
                value = "0% Participation Rate"
            return value

        ## make a content to be displayed in the picture position.
        field_picture_position = get_percent(self._participated, self._invited) + \
            "\n" + f"{self._participated:,d}" + ' / ' + f"{self._invited:,d}" + "\n" + "(Participated / Invited)"

        ## prepare image.
        img  = Image(self.input_source + "/" + self.image_src)
        img.height = 90
        img.width = 110

        ## calculate total rows that will be placed in our output file.
        total_rows = 4 + len(self._item_list)

        ## make a workbook and sheet.
        self.book = openpyxl.Workbook()
        sheet = self.book.active
        sheet.title = "{} Demographic Trends".format(self.current_year)

        ## set styles like font, color, direction, border...
        ft = Font(name="Arial", size=8)
        ft_bold = Font(name="Arial", size=8, bold=True)
        light_na_font = Font(name="Arial", size=8, color="999999")
        bold_na_font = Font(name="Arial", size=8, color="999999", bold=True)
        vertical = Alignment(textRotation=90, horizontal='center')
        center_alignment = Alignment(horizontal='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        grey_back = PatternFill("solid", fgColor="EEEEEE")
        white_back = PatternFill("solid", fgColor="FFFFFF")
        side = Side(style='thin', color="CCCCCC")
        thin_border = Border(left=side,
                     right=side,
                     top=side,
                     bottom=side)

        ## merge all needed columns and rows.
        col_num = 1
        skip_list = []
        for key in self.precious_dict:
            if len(self.precious_dict[key]) == 0:
                continue
            delta = len(self.precious_dict[key]) - 1
            if col_num == 1:
                delta += 1
            end_col = delta + col_num
            cell = sheet.cell(row=1, column=col_num)
            cell.font = ft
            cell.value = "" if key == "delta" else key
            sheet.merge_cells(start_row=1, start_column=col_num, end_row=2, end_column=end_col)
            col_num += delta + 2

            ## make a skip_list which will be used to make a empty column between groups.
            skip_list.append(col_num - 1)

        ## set empty cells white.
        for row in range(1, total_rows + 40 + 1):
            for col in range(1, col_num - 2 + 40 + 1):
                sheet.cell(row=row, column=col).fill = white_back

        ## prepare the whole data to be placed in the sheet.
        len_sub_key = 0
        frames = []
        for key in tqdm(self.precious_dict, desc="prepare the whole data"):
            sub_dict = self.precious_dict[key]
            
            ## prepare and write the first column data.
            if key == '':
                _list = []
                _list.append([field_picture_position, 1])
                _list.append(["Number of Respondents (incl. N/A)", 1])
                for criteria, item in self._item_list:
                    if criteria == 0:
                        _list.append([item, 0])
                    elif criteria == 1:
                        _list.append([self._item_pd[self._item_pd["Item ID"] == item]["Short Text [2020 onward]"].values[0], 1])
                
                ## write the first column data and set styles.
                for index, item in enumerate(_list):
                    cell = sheet.cell(row=3 + index, column=1)
                    if item[1] == 0:
                        cell.font = ft_bold
                    else:
                        cell.font = ft

                    cell.border = thin_border
                    cell.fill = grey_back
                    cell.alignment = right_alignment
                    cell.alignment = cell.alignment.copy(horizontal="left")

                    ## set value to cell.
                    try:
                        cell.value = round(Decimal(str(item[0])), 2)
                    except:
                        cell.value = item[0]

                ## autofit the first column's width.
                _ = 0
                for item in _list[1:]:
                    if len(item[0]) > _:
                        _ = len(item[0])

                sheet.column_dimensions["A"].width = _ / 1.5

                if _ < 40:
                     sheet.column_dimensions["A"].width = 26

            ## prepare the rest of data to fill other columns
            for sub_key in sub_dict:

                if len(sub_key) > len_sub_key:
                    len_sub_key = len(sub_key)

                sub_item = sub_dict[sub_key]
                _list = []
                if sub_key == "Δ Your Org ({})".format(self.past_year) and self.logic == 1:
                    _list.append(["Δ Gilead Overall ({})".format(self.past_year), 2])
                elif sub_key == "Δ Your Org ({})".format(self.past_year) and self.GM:
                    _list.append(["Δ " + self.GM + " ({})".format(self.past_year), 2])
                elif sub_key == "Δ Your Org ({})".format(self.past_year) and self.site_lead:
                    _list.append(["Δ " + self.site_lead + " ({})".format(self.past_year), 2])
                elif sub_key == "Parent Group" and self.GM and self.GM_region_human_parentorg == 0:
                    _list.append([self.GM_parent, 2])
                elif sub_key == "Parent Group":
                    _list.append([self._supervisor_last_name + " Org", 2])
                elif sub_key == "Your Org ({})".format(self.current_year) and self.GM:
                    _list.append([self.GM + " ({})".format(self.current_year), 2])
                elif sub_key == "Your Org ({})".format(self.current_year) and self.site_lead:
                    _list.append([self.site_lead + " ({})".format(self.current_year), 2])
                elif sub_key.endswith("Office"):
                    _list.append(["Office", 2])
                elif sub_key.endswith(" - COMM"):
                    _list.append([sub_key[:-7], 2])
                else:
                    _list.append([sub_key, 2])
                _list.append([self.first_row[key + sub_key], 1])

                ## prepare history column.
                if sub_key == "Δ Your Org ({})".format(self.past_year):
                    for criteria, item in self._item_list:
                        origin_item = item
                        if criteria == 1:
                            origin_item = self._item_pd[self._item_pd["Item ID"] == item]["Unique Item Code"].values[0]
                            item = self._get_past_field_name_by_current_name(origin_item)
                        try:
                            _ = sub_item[item]
                            if _[0] != "N/A":
                                _[0] = self.precious_dict[""]["Your Org ({})".format(self.current_year) if self.logic != 1 else "Gilead Overall"][origin_item][0] - _[0]
                            _list.append(_)
                        except:
                            _list.append(["N/A", 1 if criteria else 0])
                ## prepare benchmark column.
                elif sub_key == "Δ External":
                    for criteria, item in self._item_list:
                        origin_item = item
                        if criteria == 1:
                            origin_item = self._item_pd[self._item_pd["Item ID"] == item]["Unique Item Code"].values[0]
                            item = self._get_benchmark_field_name_by_current_name(origin_item)
                        try:
                            _ = sub_item[item]
                            if _[0] != "N/A":
                                _[0] = self.precious_dict[""]["Your Org ({})".format(self.current_year) if self.logic != 1 else "Gilead Overall"][origin_item][0] - _[0]
                            _list.append(_)
                        except:
                            _list.append(["N/A", 1 if criteria else 0])
                else:
                    for criteria, item in self._item_list:
                        if criteria == 1:
                            item = self._item_pd[self._item_pd["Item ID"] == item]["Unique Item Code"].values[0]

                        _list.append(sub_item[item])
                    
                frames.append(_list)
        
        ## this method is used to make a empty column when write columns.
        def get_column_number(number):

            for num in skip_list:
                if number >= num:
                    number += 1
            return number
        
        ## below method is used to calculate delta.
        def get_delta(first, second):
            return round(second * 100 - first * 100)

        gilead_org = frames[0]

        ## write the rest of the columns and set styles.
        for col_index, column in enumerate(tqdm(frames, desc="formating and styling")):
            for row_index, item in enumerate(column):
                cell = sheet.cell(row=3 + row_index, column = get_column_number(2 + col_index))

                ## set borders of all cells.
                cell.border = thin_border

                ## set background grey of the row->4.
                if row_index == 1:
                    cell.fill = grey_back
                    cell.number_format = numbers.BUILTIN_FORMATS[3]

                ## if the cell is placed in category row, set bold font style. Otherwise set general font style.
                if item[1] == 0:
                    cell.font = ft_bold
                else:
                    cell.font = ft
                cell.alignment = right_alignment

                ## set corresponding styles to row->3
                if item[1] == 2:
                    cell.alignment = center_alignment
                    cell.alignment = vertical
                    cell.fill = grey_back
                
                ## set value to cell.
                try:
                    cell.value = round(Decimal(str(item[0])), 2)
                except:
                    cell.value = item[0]

                ## make "N/A" cell lightgrey.
                if row_index >= 1:
                    if item[0] == "N/A":
                        cell.alignment = right_alignment
                        cell.font = bold_na_font if item[1] == 0 else light_na_font

                ## set background color and set format of percentage to rows below 5.
                if row_index >= 2:

                    ## set format percentage.
                    cell.number_format = numbers.FORMAT_PERCENTAGE

                    ## compare the rest of the columns with your org and set background.
                    if col_index >= self.logic:
                        try:
                            # cell.fill = PatternFill("solid", fgColor=self._get_color(get_delta(your_org[row_index][0], item[0])))
                            cell.fill = PatternFill("solid", fgColor=self._get_color(get_delta(frames[self.logic - 1][row_index][0], item[0])))
                        except:
                            ## this skip the case of N/A
                            pass

                    ## compare your org (2020) with parent org and set background.
                    elif col_index == self.logic - 1 and self.logic >= 2:
                        try:
                            cell.fill = PatternFill("solid", fgColor=self._get_color(get_delta(frames[self.logic - 2][row_index][0], item[0])))
                        except:
                            ## this skip the case of N/A
                            pass

                    ## compare parent org with gilead org and set background.
                    elif col_index == 1 and self.logic == 3:
                        try:
                            cell.fill = PatternFill("solid", fgColor=self._get_color(get_delta(gilead_org[row_index][0], item[0])))
                        except:
                            ## this skip the case of N/A
                            pass
                    
                    ## set background color about history data.
                    try:
                        _ = len(self.precious_dict["Direct Reports (as of April 24, 2018)"].keys())
                    except:
                        _ = 0
                    if col_index == self.logic + _ or col_index == self.logic + 1 + _:
                        try:
                            cell.fill = PatternFill("solid", fgColor=self._get_color(round(item[0] * 100)))
                        except:
                            ## this skip the case of N/A
                            pass
                    
                    ## force adding PLUS symbol.
                    if column[0][0][0] == "Δ":
                        try:
                            if item[0] > 0.005:
                                cell.number_format = "+0%"
                        except:
                            pass

        ## set columns' width and empty columns' width.
        for i in range(2, col_num - 2 + 1):
            name = ce.get_column_letter(i)
            if i in skip_list:
                sheet.column_dimensions[name].width = 0.8
            else:
                sheet.column_dimensions[name].width = 4.5
        
        ## set all rows' height.
        for i in range(1, total_rows + 1):
            if i == 3:
                sheet.row_dimensions[i].height = 75
                if len_sub_key > 15:
                    sheet.row_dimensions[i].height = len_sub_key * 5
            else:
                sheet.row_dimensions[i].height = 10.2

        ## freeze panes.
        sheet.freeze_panes = ce.get_column_letter(self.logic + 3) + '5'

        ## push content right in A3, A4.
        sheet['A3'].alignment = right_alignment
        sheet['A4'].alignment = right_alignment

        ## insert picture and set background white
        sheet['A3'].fill = white_back
        size = XDRPositiveSize2D(p2e(115), p2e(95))
        marker = AnchorMarker(col=0, colOff=p2e(0), row=2, rowOff=p2e(10))
        img.anchor = OneCellAnchor(_from=marker, ext=size)
        
        sheet['A3'].alignment = sheet['A3'].alignment.copy(wrapText=True, vertical="bottom")
        sheet.add_image(img)

    def _get_names_from_field(self, field_list):

        keys = [field[0] for field in field_list]
        return keys


    def _filterResource(self, id_list):

        filter_item = self._item_pd[self._item_pd["Item ID"].isin(id_list)]["Unique Item Code"].tolist()

        filter_item.insert(0, 'ExternalReference')
        self._filtered_raw_data = self.raw_data_pd[filter_item]

        ## about history data.
        self._helper_past = False
        filter_past_item = []
        for index, item in enumerate(filter_item):
            if index == 0:
                filter_past_item.append(item)
            else:
                new_item = self._get_past_field_name_by_current_name(item)
                if new_item != None:
                    filter_past_item.append(new_item)
                else:
                    self._helper_past = True
        
        self._filtered_raw_past_data = self.raw_data_past_pd[filter_past_item]

        ## about benchmark data.
        self._helper_benchmark = False
        filter_benchmark_item = []
        for index, item in enumerate(filter_item):
            if index == 0:
                filter_benchmark_item.append(item)
            else:
                new_item = self._get_benchmark_field_name_by_current_name(item)
                if new_item:
                    filter_benchmark_item.append(new_item)
                else:
                    self._helper_benchmark = True
        
        self._filtered_benchmark_value = self.benchmark_pd[self.benchmark_pd["Unique Item Code"].isin(filter_benchmark_item)]

    def _get_color(self, delta):
        if abs(delta) > 25:
            delta = delta // abs(delta) * 25
        series = self.heatmap_color_pd[self.heatmap_color_pd["Delta"] == delta]
        return str(hex(series["R"].values[0]))[2:] + str(hex(series["G"].values[0]))[2:] + str(hex(series["B"].values[0]))[2:]

    def _preProcess(self):

        ## some process about GM
        if self.GM:
            try:
                if len(self.GM_levels[self.GM_levels["GM Level 2 ID"] == self._leader_id].index) > 1:
                    self.use_affiliate = True
            except:
                pass
            try:
                if len(self.GM_levels[self.GM_levels["GM Level 3 ID"] == self._leader_id].index) > 1:
                    self.use_affiliate = True
                    self.affiliate_second = True
            except:
                pass

        ## make a item group and filter the able source.
        self._item_pd = self.item_code_pd[self.item_code_pd["Type ID"] == "T01"]

        self._benchmark_item_pd = self._item_pd[self._item_pd["External Benchmark"] == "e"]
        self._item_pd = self._item_pd[self._item_pd["External Benchmark"] == "i"]

        self._rest_item_pd = self._item_pd[~self._item_pd["Unique Item Code"].isin(self.origin_raw_data_pd.columns.values)].reset_index(drop=True)
        self._item_pd = self._item_pd[self._item_pd["Unique Item Code"].isin(self.origin_raw_data_pd.columns.values)].reset_index(drop=True)
        self._rest_item_pd = self._rest_item_pd[self._rest_item_pd["Unique Item Code"].isin(self.origin_raw_data_past_pd.columns.values)].reset_index(drop=True)
        self._both_item_pd = self._item_pd[self._item_pd["Unique Item Code"].isin(self.origin_raw_data_past_pd.columns.values)].reset_index(drop=True)

        self.category_pd = self.origin_category_pd[self.origin_category_pd["Item ID in 2020 Survey"].isin(self._item_pd["Item ID"].tolist())]
        self.category_pd = self.category_pd.drop_duplicates(subset=["Item ID in 2020 Survey"]).reset_index(drop=True)
        self.order_category = self.category_pd.drop_duplicates(subset=["2020 Category"]).loc[:, "2020 Category"].tolist()

        self._group_dict = self.category_pd.groupby(["2020 Category"]).groups

        self.raw_data_pd = self.origin_raw_data_pd.iloc[2:].reset_index(drop=True)

        ## filter the able source from history file.
        self.raw_data_past_pd = self.origin_raw_data_past_pd.iloc[2:].reset_index(drop=True)

        ## Convert numeric values into favorable or not.
        ## [1, 2, 3] -> 0, [4, 5] -> 1, [all others] -> ''
        for field in self._item_pd["Unique Item Code"]:
            new_list = []
            for item in self.raw_data_pd[field].tolist():
                if item < 4 and item > 0:
                    new_list.append(0)
                elif item >= 4 and item <= 5:
                    new_list.append(1)
                else:
                    new_list.append('')
            self.raw_data_pd[field] = new_list
            
            ## do the same process about history data.
            try:
                field = self._get_past_field_name_by_current_name(field)
                new_list_past = []
                for item in self.raw_data_past_pd[field].tolist():
                    if item < 4 and item > 0:
                        new_list_past.append(0)
                    elif item >= 4 and item <= 5:
                        new_list_past.append(1)
                    else:
                        new_list_past.append('')
                self.raw_data_past_pd[field] = new_list_past
            except:
                pass

        ## new feature -> process A/B pair.
        pairs = self._item_pd[self._item_pd["AB Code"].isin(["A", "B"])].reset_index(drop=True)
        pairs_dict = pairs.groupby(["Item ID"]).groups

        for key in pairs_dict:

            _list = pairs_dict[key]
            pairs_row = pairs.iloc[_list, :]
            item_list = pairs_row["Unique Item Code"].tolist()
            text_list = pairs_row["Short Text [2020 onward]"].tolist()
            new_text = "/".join(text_list)
            self._item_pd = self._item_pd[~(self._item_pd["Unique Item Code"] == item_list[1])].reset_index(drop=True)
            _index = self._item_pd[self._item_pd["Unique Item Code"] == item_list[0]].index
            self._item_pd.loc[_index, "Short Text [2020 onward]"] = new_text

            column_pair = self.raw_data_pd[item_list]
            _list = []
            for index in range(len(column_pair.index)):
                _row = column_pair.iloc[index, :].tolist()
                for val in _row:
                    if type(val) == type(0):
                        _list.append(val)
                        break
                else:
                    _list.append('')

            _pd = pd.DataFrame(_list, columns=[item_list[0]])
            self.raw_data_pd = self.raw_data_pd.drop(columns=[item_list[1]])
            self.raw_data_pd[item_list[0]] = _pd

        self.demographics_pd = self.origin_demographics_pd

        ## convert the demographics data to include only answered entries.
        self._invited_demographics_data = self.demographics_pd[self.demographics_pd["Invitee Flag"] == 1].reset_index(drop=True)
        self._answered_demographics_data = self._invited_demographics_data[self._invited_demographics_data.loc[:, "Worker ID"].isin(self.raw_data_pd['ExternalReference'].tolist())].reset_index(drop=True)
        self._gilead_org = self._answered_demographics_data

        ## do the same process about history data.
        self._invited_demographics_past_data = self.demographics_past_pd[self.demographics_past_pd["Invitee Flag"] == 1].reset_index(drop=True)
        self._answered_demographics_past_data = self._invited_demographics_past_data[self._invited_demographics_past_data.loc[:, "Worker ID"].isin(self.raw_data_past_pd['ExternalReference'].tolist())].reset_index(drop=True)
    
        if self.GM:
            self._gm_demographics_data = self._answered_demographics_data[self._answered_demographics_data[self.GM] == 1].reset_index(drop=True)
            self._gm_demographics_past_data = self._answered_demographics_past_data[self._answered_demographics_past_data[self.GM] == 1].reset_index(drop=True)

        if self.site_lead:
            self._site_demographics_data = self._answered_demographics_data[self._answered_demographics_data[self.site_lead] == 1].reset_index(drop=True)
            self._site_demographics_past_data = self._answered_demographics_past_data[self._answered_demographics_past_data[self.site_lead] == 1].reset_index(drop=True)      

    def _get_past_field_name_by_current_name(self, field_name):
        _item_id = self._item_pd[self._item_pd["Unique Item Code"] == field_name]["Item ID"].values[0]
        try:
            new_field_name = self._rest_item_pd[self._rest_item_pd["Item ID"] == _item_id]["Unique Item Code"].values[0]
        except:
            new_field_name = None

        if new_field_name == None:
            try:
                new_field_name = self._both_item_pd[self._both_item_pd["Unique Item Code"] == field_name]["Unique Item Code"].values[0]
            except:
                new_field_name = None

        return new_field_name

    def _get_benchmark_field_name_by_current_name(self, field_name):
        _item_id = self._item_pd[self._item_pd["Unique Item Code"] == field_name]["Item ID"].values[0]
        try:
            field_name = self._benchmark_item_pd[self._benchmark_item_pd["Item ID"] == _item_id]["Unique Item Code"].values[0]
        except:
            field_name = None
        
        return field_name

    def _prepareColumnsForID(self):
        ## find the supervisor level of the given leader id.

        self.first_row = {}
        _temp = "Supervisor Level {} ID"

        leader_entry = self.origin_demographics_pd[self.origin_demographics_pd.loc[:, "Worker ID"] == self._leader_id]

        if self._leader_id == 999999:
            self.logic = 1
            leader_level = 1
        else:
            for index, level in enumerate(leader_entry.loc[:, _temp.format(2) : _temp.format(10)]):
                if (leader_entry[level] == self._leader_id).tolist()[0]:
                    leader_level = index + 2
                    break
        
        self.output_path = "/" + leader_entry["Worker Name"].values[0]
        self.file_name = leader_entry["Worker Last Name"].values[0]

        _name = leader_entry["Worker Name"].values[0]

        supervisor_level = leader_level - 1
        direct_level = leader_level + 1

        if leader_level >= 3:
            self.logic = 3
            self._supervisor_id = leader_entry.loc[:, _temp.format(supervisor_level)].values[0]
            _supervisor_entry = self.origin_demographics_pd[self.origin_demographics_pd.loc[:, "Worker ID"] == self._supervisor_id]
            self._supervisor_last_name = _supervisor_entry["Worker Last Name"].values[0]

            ## get Parent group.
            self._parent_org = self._answered_demographics_data[self._answered_demographics_data.loc[:, _temp.format(supervisor_level)] == self._supervisor_id].reset_index(drop=True)

        elif leader_level == 2:
            self.logic = 2

        ## get Your Org data
        if self.logic == 1:
            self._your_org = self._answered_demographics_data
            self._invited = len(self._invited_demographics_data.index)
        else:
            if self.GM:
                _name = self.GM
                self._your_org = self._gm_demographics_data
                self._invited = len(self._invited_demographics_data[self._invited_demographics_data[self.GM] == 1].index)

                temp_level = "GM Level {} ID"
                ## find supervisor level.
                _corresponding_row = self.GM_levels[self.GM_levels["GM ID"] == self._leader_id]
                _parent_level = _corresponding_row["Parent Level"].values[0]

                if _parent_level == 1:
                    self.logic = 2
                else:
                    ## get parent org in case of GM and flag == 1.
                    if not bool(self.GM_region_human_parentorg):
                        _org_id = _corresponding_row[temp_level.format(_parent_level)].values[0]
                        _org_name = self.GM_levels[self.GM_levels["GM ID"] == _org_id]["GM Org"].values[0]
                        self.GM_parent = _org_name
                        
                        self._parent_org = self._answered_demographics_data[self._answered_demographics_data[_org_name] == 1].reset_index(drop=True)
            elif self.site_lead:
                _name = self.site_lead
                self._your_org = self._site_demographics_data
                self._invited = len(self._invited_demographics_data[self._invited_demographics_data[self.site_lead] == 1].index)

            else:
                self._your_org = self._answered_demographics_data[self._answered_demographics_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True)
                self._invited = len(self._invited_demographics_data[self._invited_demographics_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True).iloc[:, 0])
        
        if len(self._your_org.index) < 4:
            return self._leader_id, _name

        ## calculate nums of participated
        self._participated = len(self._your_org.loc[:, "Worker ID"])
        
        ## get your history group
        if self.GM:
            self._your_past_org = self._gm_demographics_past_data
        elif self.site_lead:
            self._your_past_org = self._site_demographics_past_data
        else:
            if self.logic == 1:
                self._your_past_org = self._answered_demographics_past_data
            else:
                self._your_past_org = self._answered_demographics_past_data[self._answered_demographics_past_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True)


        ## make direct report fields.
        if not self.GM:
            self._direct_report_field = []
            temp_org = self._your_org[~(self._your_org.loc[:, "Worker ID"] == self._leader_id)].reset_index(drop=True)
            
            if self.site_lead:
                temp_org = temp_org[temp_org.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True)

            _dict = temp_org.groupby(_temp.format(direct_level)).groups
            for key in _dict:
                try:
                    self._direct_report_field.append([self.origin_demographics_pd[self.origin_demographics_pd.loc[:, "Worker ID"] == key].loc[:, "Worker Name"].values[0], _dict[key]])
                except:
                    pass

        ## make grade group fields.
        self._grade_group_fields = []
        _dict = self._your_org.groupby("Pay Grade Group").groups
        for key in _dict:
            self._grade_group_fields.append([str(key), _dict[key]])

        ## make tenure group fields.
        self._tenure_group_fields = []
        _dict = self._your_org.groupby("Length of Service Group").groups
        for key in _dict:
            if not key == "15+ Years":
                self._tenure_group_fields.append([key, _dict[key]])
        try:
            self._tenure_group_fields.append(["15+ Years", _dict["15+ Years"]])
        except:
            pass
        
        ## make performance rating fields.
        self._performance_rating_fields = []
        _dict = self._your_org.groupby("2019 Performance Rating").groups
        _list = _dict.keys()
        for key in self.performance_order:
            try:
                self._performance_rating_fields.append([key, _dict[key]])
                _list.remove(key)
            except:
                pass
        for key in _list:
            if key == "Unspecified":
                self._performance_rating_fields.append(["No Rating", _dict[key]])
            else:
                self._performance_rating_fields.append([key, _dict[key]])

        ## make talent cordinate fields.
        self._talent_cordinate_fields = []
        _dict = self._your_org.groupby("2020 Talent Coordinate").groups
        for key in _dict:
            _key = key
            if key == "Unspecified":
                _key = "No Coordinate"
            self._talent_cordinate_fields.append([_key, _dict[key]])
        
        ## make gender fields.
        self._gender_fields = []
        _dict = self._your_org.groupby("Gender").groups
        for key in _dict:
            self._gender_fields.append([key, _dict[key]])

        ## make Ethnicity fields.
        self._ethnicity_fields = []
        _dict = self._your_org.groupby("Ethnicity (US)").groups
        for key in _dict:
            if not key == "Non-US":
                self._ethnicity_fields.append([key, _dict[key]])
        try:
            self._ethnicity_fields.append(["Non-US", _dict["Non-US"]])
        except:
            pass

        ## make age group fields.
        self._age_fields = []
        _dict = self._your_org.groupby("Age Group").groups
        for key in _dict:
            self._age_fields.append([key, _dict[key]])

        ## make country fields.
        self._country_fields = []
        _dict = self._your_org.groupby("Country").groups
        for key in _dict:
            self._country_fields.append([key, _dict[key]])

        ## make kite fields.
        self._kite_fields = []
        _dict = self._your_org.groupby("Kite Employee Flag").groups
        for key in _dict:
            self._kite_fields.append([key, _dict[key]])

        ## make office type fields.
        self._office_fields = []
        _dict = self._your_org.groupby("Office Type").groups
        for key in _dict:
            self._office_fields.append([key, _dict[key]])

        ## make Region fields.
        self._region_fields = []
        _dict = self._your_org.groupby("Location Level 2").groups
        for key in _dict:
            self._region_fields.append([key, _dict[key]])

        ## make Department fields.
        self._department_fields = []
        _dict = self._your_org.groupby("Department Level 2").groups
        for key in _dict:
            self._department_fields.append([key, _dict[key]])

        ## make Gender x Ethnicity (US) fields.
        self._gender_ethnicity_fields = []
        _dict = self._your_org.groupby(["Ethnicity (US)", "Gender"]).groups
        _new_dict = copy.deepcopy(_dict)
        for key in _dict:
            _temp = _new_dict.pop(key)
            if key[0] != "Non-US":
                _new_key = "{} {}".format(key[0], key[1])
                _new_dict.update({_new_key: _temp})
        for key in _new_dict:
            self._gender_ethnicity_fields.append([key, _new_dict[key]])

        ## make affiliate fields.
        if self.use_affiliate:
            self._affiliate_fields = []
            temp_id = "GM Level {} ID"
            temp_name = "GM Level {} Org"
            key_list = []
            def get_org_id_by_name(name):
                return self.GM_levels[self.GM_levels["GM Org"] == name]["GM ID"].values[0]
            
            if not self.affiliate_second:
                _3_levels = self.GM_levels[self.GM_levels[temp_id.format(2)] == self._leader_id]
                _3_list = _3_levels[~(_3_levels["GM ID"] == self._leader_id)][temp_name.format(3)].drop_duplicates().values
                _3_list = sorted(_3_list)
                _3_last = []
                _3_complete = []
                for item in _3_list[:]:
                    if item.startswith("Kite"):
                        _3_last.append(item)
                    else:
                        _3_complete.append(item)
                for item in _3_last:
                    _3_complete.append(item)

                for _3_name in _3_complete:
                    key_list.append(_3_name)
                    _3_id = get_org_id_by_name(_3_name)
                    _4_levels = self.GM_levels[self.GM_levels[temp_id.format(3)] == _3_id]
                    _4_list = _4_levels[~(_4_levels["GM ID"] == _3_id)][temp_name.format(4)].values
                    _4_list = sorted(_4_list)
                    _4_last = []
                    _4_complete = []
                    for item in _4_list:
                        if item.startswith("Kite"):
                            _4_last.append(item)
                        else:
                            _4_complete.append(item)
                    for item in _4_last:
                        _4_complete.append(item)

                    for _4_name in _4_complete:
                        key_list.append("    " + _4_name)
                
                for key in key_list:
                    real_key = key.strip()
                    try:
                        _list = self._your_org[self._your_org[real_key] == 1].index
                        self._affiliate_fields.append([key, _list])
                    except:
                        pass
            else:
                _4_levels = self.GM_levels[self.GM_levels[temp_id.format(3)] == self._leader_id]
                _4_list = _4_levels[~(_4_levels["GM ID"] == self._leader_id)][temp_name.format(4)].values
                _4_list = sorted(_4_list)
                _4_last = []
                _4_complete = []
                for item in _4_list:
                    if item.startswith("Kite"):
                        _4_last.append(item)
                    else:
                        _4_complete.append(item)
                for item in _4_last:
                    _4_complete.append(item)

                for _4_name in _4_complete:
                    key_list.append(_4_name)

                for key in key_list:
                    try:
                        _list = self._your_org[self._your_org[key] == 1].index
                        self._affiliate_fields.append([key, _list])
                    except:
                        pass

        self.index_match = {
            "": ["Gilead Overall", "Parent Group", "Your Org ({})".format(self.current_year)],
            "Direct Reports (as of April 24, 2018)": sorted(self._get_names_from_field(self._direct_report_field)),
            "delta": ["Δ Your Org ({})".format(self.past_year), "Δ External"],
            "Grade Group": self._get_names_from_field(self._grade_group_fields),
            "Tenure Group": self._get_names_from_field(self._tenure_group_fields),
            "Office Type": self._get_names_from_field(self._office_fields),
            "2019 Performance Rating": self._get_names_from_field(self._performance_rating_fields),
            "2020 Talent Coordinate": self._get_names_from_field(self._talent_cordinate_fields),
            "Gender": self._get_names_from_field(self._gender_fields),
            "Ethnicity (US)": self._get_names_from_field(self._ethnicity_fields),
            "Gender x Ethnicity (US)": self._get_names_from_field(self._gender_ethnicity_fields),
            "Age Group": self._get_names_from_field(self._age_fields),
            "Function": self._get_names_from_field(self._department_fields),
            "Region": self._get_names_from_field(self._region_fields),
            "Country": self._get_names_from_field(self._country_fields),
            "Kite": self._get_names_from_field(self._kite_fields),
        }

        if self.logic == 2:
            self.index_match.update({"": ["Gilead Overall", "Your Org ({})".format(self.current_year)]})
        elif self.logic == 1:
            self.index_match.update({"": ["Gilead Overall"]})

        if self.GM:
            self.index_match.pop("Direct Reports (as of April 24, 2018)", "no")

        if self.use_affiliate:
            self.index_match.pop("Country", "no")
            _keys = list(self.index_match.keys())
            _keys.insert(_keys.index("delta") + 1, "Affiliate")
            _dict = self.index_match
            self.index_match = {}
            for key in _keys:
                if key == "Affiliate":
                    self.index_match[key] = self._get_names_from_field(self._affiliate_fields)
                else:
                    self.index_match[key] = _dict[key]

        ## make and initiate a dict to save all calculated data.
        self.precious_dict = {}
        for first_index in self.index_match:
            _ = {}
            for item in self.index_match[first_index]:
                _.update({item: {}})
            self.precious_dict.update({first_index: _})

    def _calculateEachRow(self, item):
        
        ## calculate Gilead overall %s
        _dict, lens = self._calculateOverall(self._gilead_org, item)
        self.precious_dict[""]["Gilead Overall"].update(_dict)
        self.first_row.update({"Gilead Overall": lens})

        ## calculate Parent Group %s
        if self.logic == 3:
            _dict, lens = self._calculateOverall(self._parent_org, item)
            self.precious_dict[""]["Parent Group"].update(_dict)
            self.first_row.update({"Parent Group": lens})

        ## calculate Your Org(2018) %s
        if self.logic >= 2:
            _dict, lens = self._calculateOverall(self._your_org, item)
            self.precious_dict[""]["Your Org ({})".format(self.current_year)].update(_dict)
            self.first_row.update({"Your Org ({})".format(self.current_year): lens})

        ## calculate Δ Your Org (2018) %s
        if len(self._filtered_raw_past_data.columns) > 1:
            _dict, lens = self._calculateOverall(self._your_past_org, item, history=True)
            self.precious_dict["delta"]["Δ Your Org ({})".format(self.past_year)].update(_dict)
            self.first_row.update({"deltaΔ Your Org ({})".format(self.past_year): lens})

        ## calculate Δ External %s
        if len(self._filtered_benchmark_value.columns) > 1:
            self.first_row.update({"deltaΔ External": "N/A"})
            _dict = {}
            for index, row in self._filtered_benchmark_value.iterrows():
                value = row["External - CAmp Biotechnology & Medical Devices 2019"]
                if not value:
                    _dict.update({row["Unique Item Code"]: [value, 1]})
                else:
                    _dict.update({row["Unique Item Code"]: [value, 1]})
            _dict.update({item: ["N/A", 0]})
            self.precious_dict["delta"]["Δ External"].update(_dict)

        ## calculate Direct reports %s
        if not self.GM:
            self._calculateSubFields(self._direct_report_field, "Direct Reports (as of April 24, 2018)", item)
        
        ## calcualte Grade Group %s
        self._calculateSubFields(self._grade_group_fields, "Grade Group", item)

        ## calcualte Tenure Group %s
        self._calculateSubFields(self._tenure_group_fields, "Tenure Group", item)

        ## calculate Performance Rating %s
        self._calculateSubFields(self._performance_rating_fields, "2019 Performance Rating", item)

        ## calculate Talent Coordinate %s
        self._calculateSubFields(self._talent_cordinate_fields, "2020 Talent Coordinate", item)

        ## calculate Gender %s
        self._calculateSubFields(self._gender_fields, "Gender", item)

        ## calculate Ethnicity (US) %s
        self._calculateSubFields(self._ethnicity_fields, "Ethnicity (US)", item)

        ## calculate Age Group %s
        self._calculateSubFields(self._age_fields, "Age Group", item)

        ## calculate Country %s
        if not self.use_affiliate:
            self._calculateSubFields(self._country_fields, "Country", item)

        ## calculate Kite %s
        self._calculateSubFields(self._kite_fields, "Kite", item)

        ## calculate office %s
        self._calculateSubFields(self._office_fields, "Office Type", item)

        ## calculate regions %s
        self._calculateSubFields(self._region_fields, "Region", item)

        ## calculate department %s
        self._calculateSubFields(self._department_fields, "Function", item)

        ## calculate affiliate %s
        if self.use_affiliate:
            self._calculateSubFields(self._affiliate_fields, "Affiliate", item)

        ## calculate gender ethnicity %s
        self._calculateSubFields(self._gender_ethnicity_fields, "Gender x Ethnicity (US)", item)

    def _calculateOverall(self, dataframe, item, history=False):

        ## calcualte overall fields.
        ids = dataframe.iloc[:, 0]
        # nums = len(ids)
        if history:
            working_pd = self._filtered_raw_past_data[self._filtered_raw_past_data["ExternalReference"].isin(ids)].reset_index(drop=True)
        else:
            working_pd = self._filtered_raw_data[self._filtered_raw_data["ExternalReference"].isin(ids)].reset_index(drop=True)
        nums = len(working_pd.iloc[:, 0])
        return self._get_sum(working_pd.iloc[:, 1:], nums, item, history), len(ids)
    
    def _calculateSubFields(self, dataframe, column_name, item):

        ## calculate the sum of all sub fields except overall fields.
        for field in dataframe:
            dictionary = self.precious_dict[column_name]
            _list = field[1]
            ids = self._your_org.iloc[_list, 0].tolist()
            nums = len(ids)
            working_pd = self._filtered_raw_data[self._filtered_raw_data["ExternalReference"].isin(ids)].reset_index(drop=True)
            _dict = self._get_sum(working_pd.iloc[:, 1:], nums, item)
            dictionary[field[0]].update(_dict)
            self.first_row.update({column_name + field[0]: len(ids)})
    
    def _get_sum(self, data, nums, item, implicity=False):

        ## calculate the sum of favorable scores.
        _dict = {}
        determine_parent_na = False

        for ind in range(len(data.columns)):
            _ = data.iloc[:, ind]
            is_empty_column = True
            lens = nums
            sub = 0
            count_valid = 0
            for __ in _:
                if __ != '':
                    is_empty_column = False
                    sub += __
                    count_valid += 1
                else:
                    lens -= 1
            if is_empty_column:
                _dict.update({data.columns.values[ind]: ["N/A", 1]})
                determine_parent_na = True
            else:
                if count_valid >= 4:
                    _dict.update({data.columns.values[ind]: [sub / lens, 1]})
                else:
                    _dict.update({data.columns.values[ind]: ["N/A", 1]})
                    determine_parent_na = True
        
        if determine_parent_na:
            _dict.update({item: ["N/A", 0]})
        else:
            total_lens = nums
            total = 0
            cols = len(data.columns)
            for ind in range(nums):
                _series = data.iloc[ind, :]
                sub_sum = 0
                _is_nan = False
                for val in _series:
                    if val != '':
                        sub_sum += val
                    else:
                        _is_nan = True
                        total_lens -= 1
                        break

                if not _is_nan:
                    total += sub_sum / cols

            if total_lens < 4:
                _dict.update({item: ["N/A", 0]})
            else:
                _dict.update({item: [total / total_lens, 0]})

        if implicity:
                _dict.update({item: ["N/A", 0]})
        return _dict

class LTMaker:

    def __init__(self, **args):
        ## init method
        self.image_src = args['image']

        self.raw_data_file = args['raw_data']
        self.raw_data_past_file = args['raw_data_past']
        self.item_code_file = args['item_code']
        self.demographics_file = args['demographics']
        self.demographics_past_file = args['demographics_past']
        self.heatmap_color_file = args['heatmap_color']
        self.leader_file = args['leader']
        self.output_source = args['output_folder']
        self.input_source = args['input_folder']
        self.how2use_file = args['how to use']
        self.GM_levels_file = args['gm_levels']

        self.performance_order = [
            "Exceptional",
            "Exceeded",
            "Achieved",
            "Improvement Needed",
            "On Leave",
            "No Rating",
        ]

        self.current_year = self.demographics_file[:4]
        self.past_year = self.demographics_past_file[:4]

        self.GM_region_human_parentorg = 0

    def readAllFiles(self):
        ## read all needed files to make report.
        self.origin_raw_data_pd = pd.read_excel(self.input_source + "/" + self.raw_data_file, engine="openpyxl")
        self.origin_raw_data_past_pd = pd.read_excel(self.input_source + "/" + self.raw_data_past_file, engine="openpyxl")
        self.item_code_pd = pd.read_excel(self.input_source + "/" + self.item_code_file, engine="openpyxl", sheet_name="ItemCodeSTAR")
        self.origin_category_pd =  pd.read_excel(self.input_source + "/" + self.item_code_file, engine="openpyxl", sheet_name="CurrentCategorySTAR")
        self.origin_demographics_pd = pd.read_excel(self.input_source + "/" + self.demographics_file, engine="openpyxl")
        self.demographics_past_pd = pd.read_excel(self.input_source + "/" + self.demographics_past_file, engine="openpyxl")
        self.heatmap_color_pd = pd.read_excel(self.input_source + "/" + self.heatmap_color_file, engine="openpyxl")
        self.leaders = pd.read_excel(self.input_source + "/" + self.leader_file, engine="openpyxl", sheet_name="Leader")
        self.site_leads = pd.read_excel(self.input_source + "/" + self.leader_file, engine="openpyxl", sheet_name="Site Leader")
        self.GMs = pd.read_excel(self.input_source + "/" + self.leader_file, engine="openpyxl", sheet_name="GM")
        self.how2use_pd = pd.read_excel(self.input_source + "/" + self.how2use_file, engine="openpyxl", sheet_name="Score Details How to Use")
        self.GM_levels = pd.read_excel(self.input_source + "/" + self.GM_levels_file, engine="openpyxl")

    def setLeader(self, id, GM=False, site_lead=False):
        self.use_affiliate = False
        self.affiliate_second = False

        self._leader_id = id
        self.GM = GM
        self.site_lead = site_lead

    def setGMParentFlag(self, value):
        self.GM_region_human_parentorg = value

    def calculateValues(self):
        ## do some process referred to individual leader ID.
        self._item_list = []
        self._preProcess()
        self._prepareColumnsForID()

        item_list = []
        item_dict = {}
        category_list = []

        ## fill the above variables.
        for key in self.order_category:
            category_list.append(key)
            item_list.append([0, key])
            temp_list = []
            for item in self.category_pd.iloc[self._group_dict[key], 0]:
                item_list.append([1, item])
                temp_list.append(item)
            item_dict.update({key:temp_list})

        ## do main calculation process by iterating over the category list.
        for category in tqdm(category_list, desc="iterating over the list of category"):
            sub_id_list = item_dict[category]

            ## filter the source data frame to have columns belonged to each category.
            self._filterResource(sub_id_list, category)

            ## based on the filtered source, calculate the values for each row.
            self._calculateEachRow(category)

    def setWorkBook(self, book):
        self.book = book

    def makeReport(self):

        ## eliminate columns of which respondents are lower than 4.
        criteria_dict = copy.deepcopy(self.precious_dict)
        for key in criteria_dict:
            if key != "":
                for sub_key in criteria_dict[key]:
                    try:
                        if self.first_row["current"][key + sub_key] < 4 or self.first_row["past"][key + sub_key] < 4:
                            del self.precious_dict[key][sub_key]
                    except:
                        del self.precious_dict[key][sub_key]

        ## eliminate section of which segments are lower than 2.
        criteria_dict = copy.deepcopy(self.precious_dict)
        for key in criteria_dict:
            if len(criteria_dict[key]) < 2 and key != "":
                del self.precious_dict[key]

        def get_percent(numerator, denominator):
            try:
                value = str(round(numerator / denominator * 100)) + "% \n"
            except:
                value = "0% \n"
            return value

        ## make a content to be displayed in the picture position.
        field_picture_position = self.current_year + " Participation Rate " + get_percent(self._participated, self._invited) + \
            f"{self._participated:,d}" + ' / ' + f"{self._invited:,d}" + "\n" + self.past_year + " Participation Rate " + \
                get_percent(self._participated_past, self._invited_past) + f"{self._participated_past:,d}" + ' / ' + f"{self._invited_past:,d}"

        ## calculate total rows that will be placed in our output file.
        total_rows = 5 + len(self._item_list)

        ## make a workbook and sheet.
        # self.book = openpyxl.Workbook()
        sheet = self.book.create_sheet(self.current_year + "vs." + self.past_year + " Longitudinal Trends")
        # sheet.title = self.current_year + "vs." + self.past_year + " Longitudinal Trends"

        ## set styles like font, color, direction, border...
        ft = Font(name="Arial", size=8)
        ft_bold = Font(name="Arial", size=8, bold=True)
        light_na_font = Font(name="Arial", size=8, color="999999")
        bold_na_font = Font(name="Arial", size=8, color="999999", bold=True)
        vertical = Alignment(textRotation=90, horizontal='center')
        center_alignment = Alignment(horizontal='center')
        right_alignment = Alignment(horizontal='right', vertical='center')
        grey_back = PatternFill("solid", fgColor="EEEEEE")
        white_back = PatternFill("solid", fgColor="FFFFFF")
        side = Side(style='thin', color="CCCCCC")
        thin_border = Border(left=side,
            right=side,
            top=side,
            bottom=side)

        ## merge all needed columns and rows.
        col_num = 1
        skip_list = []
        for key in self.precious_dict:
            if len(self.precious_dict[key]) == 0:
                continue
            delta = len(self.precious_dict[key]) - 1
            if col_num == 1:
                delta += 1
            end_col = delta + col_num
            cell = sheet.cell(row=1, column=col_num)
            cell.font = ft
            cell.value = "" if key == "delta" else key
            sheet.merge_cells(start_row=1, start_column=col_num, end_row=2, end_column=end_col)
            col_num += delta + 2

            ## make a skip_list which will be used to make a empty column between groups.
            skip_list.append(col_num - 1)
        
        ## set empty cells white.
        for row in range(1, total_rows + 40 + 1):
            for col in range(1, col_num - 2 + 40 + 1):
                sheet.cell(row=row, column=col).fill = white_back

        ## prepare the whole data to be placed in the sheet.
        frames = []
        len_sub_key = 0
        for key in tqdm(self.precious_dict, desc="prepare the whole data"):
            sub_dict = self.precious_dict[key]

            ## prepare and write the first column data.
            if key == '':
                _list = []
                _list.append([field_picture_position, 1])
                _list.append([self.current_year + " Number of Respondents (incl. N/A)", 1])
                _list.append([self.past_year + " Number of Respondents (incl. N/A)", 1])
                for criteria, item in self._item_list:
                    try:
                        if criteria == 0:
                            _list.append([item, 0])
                        elif criteria == 1:
                            _list.append([self._item_pd[self._item_pd["Unique Item Code"] == item]["Short Text [2020 onward]"].values[0], 1])
                    except:
                        pass

                ## write the first column data and set styles.
                for index, item in enumerate(_list):
                    cell = sheet.cell(row=3 + index, column=1)
                    if item[1] == 0:
                        cell.font = ft_bold
                    else:
                        cell.font = ft

                    cell.border = thin_border
                    cell.fill = grey_back
                    cell.alignment = right_alignment
                    cell.alignment = cell.alignment.copy(horizontal="left")

                    ## set value to cell.
                    try:
                        cell.value = round(Decimal(str(item[0])), 2)
                    except:
                        cell.value = item[0]
                
                ## autofit the first column's width.
                _ = 0
                for item in _list[1:]:
                    if len(item[0]) > _:
                        _ = len(item[0])

                sheet.column_dimensions["A"].width = _ / 1.5

                if _ < 40:
                     sheet.column_dimensions["A"].width = 26

            ## prepare the rest of data to fill other columns
            for sub_key in sub_dict:

                if len(str(sub_key)) > len_sub_key:
                    len_sub_key = len(sub_key)

                sub_item = sub_dict[sub_key]
                _list = []
                if sub_key == "Parent Group Delta" and self.GM and self.GM_region_human_parentorg == 0:
                    _list.append([self.GM_parent + " Delta", 2])
                elif sub_key == "Parent Group Delta":
                    _list.append([self._supervisor_last_name + " Org Delta", 2])
                elif sub_key == "Your Org Delta ({} to {})".format(self.current_year, self.past_year) and self.GM:
                    _list.append([self.GM + " Delta ({} to {})".format(self.current_year, self.past_year), 2])
                elif sub_key == "Your Org Delta ({} to {})".format(self.current_year, self.past_year) and self.site_lead:
                    _list.append([self.site_lead + " Delta ({} to {})".format(self.current_year, self.past_year), 2])
                elif sub_key.endswith("Office"):
                    _list.append(["Office", 2])
                elif sub_key.endswith(" - COMM"):
                    _list.append([sub_key[:-7], 2])
                else:
                    _list.append([sub_key, 2])

                try:
                    _list.append([self.first_row["current"][key + sub_key], 1])
                    _list.append([self.first_row["past"][key + sub_key], 1])
                except:
                    pass

                for criteria, item in self._item_list:
                    
                    try:
                        _list.append(sub_item[item])
                    except:
                        pass

                frames.append(_list)

        ## this method is used to make a empty column when write columns.
        def get_column_number(number):

            for num in skip_list:
                if number >= num:
                    number += 1
            return number

        ## set columns' width and empty columns' width.
        for i in range(2, col_num - 2 + 1):
            name = ce.get_column_letter(i)
            if i in skip_list:
                sheet.column_dimensions[name].width = 0.8
            else:
                sheet.column_dimensions[name].width = 4.5

        ## write the rest of the columns and set styles.
        for col_index, column in enumerate(tqdm(frames, desc="formating and styling")):
            for row_index, item in enumerate(column):
                cell = sheet.cell(row=3 + row_index, column = get_column_number(2 + col_index))

                ## set borders of all cells.
                cell.border = thin_border

                ## set background grey of the row->4.
                if row_index <= 2:
                    cell.fill = grey_back
                    cell.number_format = numbers.BUILTIN_FORMATS[3]

                ## if the cell is placed in category row, set bold font style. Otherwise set general font style.
                if item[1] == 0:
                    cell.font = ft_bold
                else:
                    cell.font = ft
                    
                    ## set value to cell.
                    try:
                        cell.value = round(Decimal(str(item[0])), 2)
                    except:
                        cell.value = item[0]

                cell.alignment = right_alignment

                ## set corresponding styles to row->3
                if item[1] == 2:
                    cell.alignment = center_alignment
                    cell.alignment = vertical
                    cell.fill = grey_back

                ## make "N/A" cell lightgrey.
                if row_index >= 1:
                    if item[0] == "N/A":
                        cell.alignment = right_alignment
                        cell.font = bold_na_font if item[1] == 0 else light_na_font

                ## set background color and set format of percentage to rows below 6.
                if row_index >= 3:

                    ## set format percentage.
                    cell.number_format = numbers.FORMAT_PERCENTAGE

                    try:
                        cell.fill = PatternFill("solid", fgColor=self._get_color(round(item[0] * 100)))
                    except:
                        pass

                    try:
                        if abs(round(item[0] * 100)) == 100:
                            sheet.column_dimensions[ce.get_column_letter(get_column_number(2 + col_index))].width = 5.2
                    except:
                        pass

                    try:
                        if item[0] > 0.005:
                            cell.number_format = "+0%"
                    except:
                        pass

        ## set all rows' height.
        for i in range(1, total_rows + 1):
            if i == 3:
                sheet.row_dimensions[i].height = 75
                if len_sub_key > 15:
                    sheet.row_dimensions[i].height = len_sub_key * 5
            else:
                sheet.row_dimensions[i].height = 10.2
        
        ## freeze panes.
        sheet.freeze_panes = ce.get_column_letter(self.logic + 3) + '6'

        ## push content right in A3, A4, A5.
        sheet['A3'].alignment = right_alignment
        sheet['A4'].alignment = right_alignment
        sheet['A5'].alignment = right_alignment

        ## insert picture and set background white
        sheet['A3'].fill = white_back
        # size = XDRPositiveSize2D(p2e(115), p2e(95))
        # marker = AnchorMarker(col=0, colOff=p2e(0), row=2, rowOff=p2e(10))
        # img.anchor = OneCellAnchor(_from=marker, ext=size)

        sheet['A3'].alignment = sheet['A3'].alignment.copy(wrapText=True, vertical="bottom")
        # sheet.add_image(img)

        ## insert how to use sheet.
        how_sheet = self.book.create_sheet("How to Use")
        how_font = Font(name="Arial", size=10)
        cell = how_sheet.cell(column=1, row=1)
        cell.value = self.how2use_pd.columns.values[0]
        cell.alignment = Alignment(wrapText=True)
        cell.font = how_font
        how_sheet.row_dimensions[1].height = int(len(self.how2use_pd.columns.values[0]) / 120 * (40 / 3))
        

        for index in range(len(self.how2use_pd.index)):
            _ = 0
            for col_index in range(len(self.how2use_pd.columns)):
                cell = how_sheet.cell(column=1 + col_index, row=2 + index)

                content = self.how2use_pd.iloc[index, col_index]
                if _ < len(content):
                    _ = len(content)
                cell.value = content
                cell.alignment = Alignment(wrapText=True)
                cell.font = how_font
                how_sheet.column_dimensions[ce.get_column_letter(col_index + 1)].width = 135

            how_sheet.row_dimensions[index + 2].height = int(_ / 110 * (40 / 3))


        for row in range(1, 41 + len(self.how2use_pd.index)):
            for col in range(1, 41 + len(self.how2use_pd.columns)):
                how_sheet.cell(row=row, column=col).fill = white_back

    def writeOutput(self):

        ## specify the path of output file
        if self.GM:
            self.file_name = self.GM[:-4]
            self.output_path = "/" + self.file_name
        if self.site_lead:
            self.file_name = self.site_lead.replace(" / ", " ")
            self.output_path = "/" + self.file_name
        # path = self.output_source + self.output_path + "/" + "2021-04 Global Employee Survey - " + self.file_name + " - Longitudinal Trends.xlsx"
        path = self.output_source + self.output_path + "/" + "2021-04 Global Employee Survey - " + self.file_name + " - Score Details.xlsx"

        ## if the output file already exists, remove it.
        if os.path.exists(path):
            os.remove(path)

        ## make a folder to involve the output file.
        os.makedirs(self.output_source + self.output_path, exist_ok=True)

        ## and write the output file.
        self.book.save(path)

    def _preProcess(self):
        ## some process about GM
        if self.GM:
            try:
                if len(self.GM_levels[self.GM_levels["GM Level 2 ID"] == self._leader_id].index) > 1:
                    self.use_affiliate = True
            except:
                pass
            try:
                if len(self.GM_levels[self.GM_levels["GM Level 3 ID"] == self._leader_id].index) > 1:
                    self.use_affiliate = True
                    self.affiliate_second = True
            except:
                pass

        ## make a item group and filter the able source.
        self._item_pd = self.item_code_pd[self.item_code_pd["Type ID"] == "T01"]

        self._item_pd = self._item_pd[self._item_pd["External Benchmark"] == "i"]

        self._rest_item_pd = self._item_pd[~self._item_pd["Unique Item Code"].isin(self.origin_raw_data_pd.columns.values)].reset_index(drop=True)
        self._item_pd = self._item_pd[self._item_pd["Unique Item Code"].isin(self.origin_raw_data_pd.columns.values)].reset_index(drop=True)
        self._rest_item_pd = self._rest_item_pd[self._rest_item_pd["Unique Item Code"].isin(self.origin_raw_data_past_pd.columns.values)].reset_index(drop=True)
        self._both_item_pd = self._item_pd[self._item_pd["Unique Item Code"].isin(self.origin_raw_data_past_pd.columns.values)].reset_index(drop=True)
        
        self.category_pd = self.origin_category_pd[self.origin_category_pd["Item ID in 2020 Survey"].isin(self._item_pd["Item ID"].tolist())]
        self.category_pd = self.category_pd.drop_duplicates(subset=["Item ID in 2020 Survey"]).reset_index(drop=True)
        self.order_category = self.category_pd.drop_duplicates(subset=["2020 Category"]).loc[:, "2020 Category"].tolist()

        self._group_dict = self.category_pd.groupby(["2020 Category"]).groups

        self.raw_data_pd = self.origin_raw_data_pd.iloc[2:].reset_index(drop=True)

        ## filter the able source from history file.
        self.raw_data_past_pd = self.origin_raw_data_past_pd.iloc[2:].reset_index(drop=True)

        ## Convert numeric values into favorable or not.
        ## [1, 2, 3] -> 0, [4, 5] -> 1, [all others] -> ''
        for field in self._item_pd["Unique Item Code"]:
            new_list = []
            for item in self.raw_data_pd[field].tolist():
                if item < 4 and item > 0:
                    new_list.append(0)
                elif item >= 4 and item <= 5:
                    new_list.append(1)
                else:
                    new_list.append('')
            self.raw_data_pd[field] = new_list
                        
                ## do the same process about history data.
            try:
                field = self._get_past_field_name_by_current_name(field)
                new_list_past = []
                for item in self.raw_data_past_pd[field].tolist():
                    if item < 4 and item > 0:
                        new_list_past.append(0)
                    elif item >= 4 and item <= 5:
                        new_list_past.append(1)
                    else:
                        new_list_past.append('')
                self.raw_data_past_pd[field] = new_list_past
            except:
                pass
        ## new feature -> process A/B pair.
        pairs = self._item_pd[self._item_pd["AB Code"].isin(["A", "B"])].reset_index(drop=True)
        pairs_dict = pairs.groupby(["Item ID"]).groups

        for key in pairs_dict:

            _list = pairs_dict[key]
            pairs_row = pairs.iloc[_list, :]
            item_list = pairs_row["Unique Item Code"].tolist()
            text_list = pairs_row["Short Text [2020 onward]"].tolist()
            new_text = "/".join(text_list)
            self._item_pd = self._item_pd[~(self._item_pd["Unique Item Code"] == item_list[1])].reset_index(drop=True)
            _index = self._item_pd[self._item_pd["Unique Item Code"] == item_list[0]].index
            self._item_pd.loc[_index, "Short Text [2020 onward]"] = new_text

            column_pair = self.raw_data_pd[item_list]
            _list = []
            for index in range(len(column_pair.index)):
                _row = column_pair.iloc[index, :].tolist()
                for val in _row:
                    if type(val) == type(0):
                        _list.append(val)
                        break
                else:
                        _list.append('')

            _pd = pd.DataFrame(_list, columns=[item_list[0]])
            self.raw_data_pd = self.raw_data_pd.drop(columns=[item_list[1]])
            self.raw_data_pd[item_list[0]] = _pd

        self.demographics_pd = self.origin_demographics_pd
        
        ## convert the demographics data to include only answered entries.
        self._invited_demographics_data = self.demographics_pd[self.demographics_pd["Invitee Flag"] == 1].reset_index(drop=True)
        self._answered_demographics_data = self._invited_demographics_data[self._invited_demographics_data.loc[:, "Worker ID"].isin(self.raw_data_pd['ExternalReference'].tolist())].reset_index(drop=True)
        self._gilead_org = self._answered_demographics_data

        ## do the same process about history data.
        self._invited_demographics_past_data = self.demographics_past_pd[self.demographics_past_pd["Invitee Flag"] == 1].reset_index(drop=True)
        self._answered_demographics_past_data = self._invited_demographics_past_data[self._invited_demographics_past_data.loc[:, "Worker ID"].isin(self.raw_data_past_pd['ExternalReference'].tolist())].reset_index(drop=True)
        self._gilead_past_org = self._answered_demographics_past_data

        if self.GM:
            self._gm_demographics_data = self._answered_demographics_data[self._answered_demographics_data[self.GM] == 1].reset_index(drop=True)
            self._gm_demographics_past_data = self._answered_demographics_past_data[self._answered_demographics_past_data[self.GM] == 1].reset_index(drop=True)

        if self.site_lead:
            self._site_demographics_data = self._answered_demographics_data[self._answered_demographics_data[self.site_lead] == 1].reset_index(drop=True)
            self._site_demographics_past_data = self._answered_demographics_past_data[self._answered_demographics_past_data[self.site_lead] == 1].reset_index(drop=True)      

    def _get_past_field_name_by_current_name(self, field_name):
        _item_id = self._item_pd[self._item_pd["Unique Item Code"] == field_name]["Item ID"].values[0]
        try:
            new_field_name = self._rest_item_pd[self._rest_item_pd["Item ID"] == _item_id]["Unique Item Code"].values[0]
        except:
            new_field_name = None

        if new_field_name == None:
            try:
                new_field_name = self._both_item_pd[self._both_item_pd["Unique Item Code"] == field_name]["Unique Item Code"].values[0]
            except:
                new_field_name = None

        return new_field_name

    def _prepareColumnsForID(self):
        ## find the supervisor level of the given leader id.
        self.first_row = {"current": {}, "past": {}}
        _temp = "Supervisor Level {} ID"

        leader_entry = self.origin_demographics_pd[self.origin_demographics_pd.loc[:, "Worker ID"] == self._leader_id]

        if self._leader_id == 999999:
            self.logic = 1
            leader_level = 1
        else:
            for index, level in enumerate(leader_entry.loc[:, _temp.format(2) : _temp.format(10)]):
                if (leader_entry[level] == self._leader_id).tolist()[0]:
                    leader_level = index + 2
                    break

        self.output_path = "/" + leader_entry["Worker Name"].values[0]
        self.file_name = leader_entry["Worker Last Name"].values[0]

        supervisor_level = leader_level - 1
        direct_level = leader_level + 1

        if leader_level >= 3:
            self.logic = 3
            self._supervisor_id = leader_entry.loc[:, _temp.format(supervisor_level)].values[0]
            _supervisor_entry = self.origin_demographics_pd[self.origin_demographics_pd.loc[:, "Worker ID"] == self._supervisor_id]
            self._supervisor_last_name = _supervisor_entry["Worker Last Name"].values[0]

            ## get Parent group.
            self._parent_org = self._answered_demographics_data[self._answered_demographics_data.loc[:, _temp.format(supervisor_level)] == self._supervisor_id].reset_index(drop=True)
            self._parent_past_org = self._answered_demographics_past_data[self._answered_demographics_past_data.loc[:, _temp.format(supervisor_level)] == self._supervisor_id].reset_index(drop=True)

        elif leader_level == 2:
            self.logic = 2

        ## get Your Org data
        if self.logic == 1:
            self._your_org = self._answered_demographics_data
            self._invited = len(self._invited_demographics_data.index)
            self._invited_past = len(self._invited_demographics_past_data.index)
        else:
            if self.GM:
                self._your_org = self._gm_demographics_data
                self._invited = len(self._invited_demographics_data[self._invited_demographics_data[self.GM] == 1].index)
                self._invited_past = len(self._invited_demographics_past_data[self._invited_demographics_past_data[self.GM] == 1].index)
                
                temp_level = "GM Level {} ID"
                ## find supervisor level.
                _corresponding_row = self.GM_levels[self.GM_levels["GM ID"] == self._leader_id]
                _parent_level = _corresponding_row["Parent Level"].values[0]

                if _parent_level == 1:
                    self.logic = 2
                else:
                    ## get parent org in case of GM and flag == 1.
                    if not bool(self.GM_region_human_parentorg):
                        _org_id = _corresponding_row[temp_level.format(_parent_level)].values[0]
                        _org_name = self.GM_levels[self.GM_levels["GM ID"] == _org_id]["GM Org"].values[0]
                        self.GM_parent = _org_name
                        
                        self._parent_org = self._answered_demographics_data[self._answered_demographics_data[_org_name] == 1].reset_index(drop=True)
                        self._parent_past_org = self._answered_demographics_past_data[self._answered_demographics_past_data[_org_name] == 1].reset_index(drop=True)
            elif self.site_lead:
                self._your_org = self._site_demographics_data
                self._invited = len(self._invited_demographics_data[self._invited_demographics_data[self.site_lead] == 1].index)
                self._invited_past = len(self._invited_demographics_past_data[self._invited_demographics_past_data[self.site_lead] == 1].index)
            
            else:
                self._your_org = self._answered_demographics_data[self._answered_demographics_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True)
                self._invited = len(self._invited_demographics_data[self._invited_demographics_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True).iloc[:, 0])
                self._invited_past = len(self._invited_demographics_past_data[self._invited_demographics_past_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True).iloc[:, 0])

        ## get your history group
        if self.GM:
            self._your_past_org = self._gm_demographics_past_data
        elif self.site_lead:
            self._your_past_org = self._site_demographics_past_data
        else:
            if self.logic == 1:
                self._your_past_org = self._answered_demographics_past_data
            else:
                self._your_past_org = self._answered_demographics_past_data[self._answered_demographics_past_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True)


        ## calculate nums of participated
        self._participated = len(self._your_org.loc[:, "Worker ID"])
        self._participated_past = len(self._your_past_org.loc[:, "Worker ID"])

        ## make direct report fields.
        if not self.GM:
            self._direct_report_field = []
            self.temp_org = self._your_org[~(self._your_org.loc[:, "Worker ID"] == self._leader_id)].reset_index(drop=True)

            if self.site_lead:
                self.temp_org = self.temp_org[self.temp_org.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True)

            _dict = self.temp_org.groupby(_temp.format(direct_level)).groups
            for key in _dict:
                try:
                    self._direct_report_field.append([self.origin_demographics_pd[self.origin_demographics_pd.loc[:, "Worker ID"] == key].loc[:, "Worker Name"].values[0], _dict[key]])
                except:
                    pass
            
            self._direct_report_past_field = {}
            self.temp_past_org = self._your_past_org[~(self._your_past_org.loc[:, "Worker ID"] == self._leader_id)].reset_index(drop=True)
            
            if self.site_lead:
                self.temp_past_org = self.temp_past_org[self.temp_past_org.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True)

            _dict_past = self.temp_past_org.groupby(_temp.format(direct_level)).groups
            for key in _dict:
                try:
                    _values = _dict_past[key]
                except:
                    _values = [0]
                self._direct_report_past_field.update({self.origin_demographics_pd[self.origin_demographics_pd.loc[:, "Worker ID"] == key].loc[:, "Worker Name"].values[0]: _values})

        ## make grade group fields.
        self._grade_group_fields = []
        _dict = self._your_org.groupby("Pay Grade Group").groups
        for key in _dict:
            self._grade_group_fields.append([str(key), _dict[key]])

        self._grade_group_past_fields = {}
        _dict = self._your_past_org.groupby("Pay Grade Group").groups
        for key in _dict:
            self._grade_group_past_fields.update({str(key): _dict[key]})

        ## make tenure group fields.
        self._tenure_group_fields = []
        _dict = self._your_org.groupby("Length of Service Group").groups
        for key in _dict:
            if not key == "15+ Years":
                self._tenure_group_fields.append([key, _dict[key]])
        try:
            self._tenure_group_fields.append(["15+ Years", _dict["15+ Years"]])
        except:
            pass

        self._tenure_group_past_fields = {}
        _dict = self._your_past_org.groupby("Length of Service Group").groups
        for key in _dict:
            self._tenure_group_past_fields.update({key: _dict[key]})

        ## make performance rating fields.
        self._performance_rating_fields = []
        _dict = self._your_org.groupby("2019 Performance Rating").groups
        _list = _dict.keys()
        for key in self.performance_order:
            try:
                self._performance_rating_fields.append([key, _dict[key]])
                _list.remove(key)
            except:
                pass
        for key in _list:
            if key == "Unspecified":
                self._performance_rating_fields.append(["No Rating", _dict[key]])
            else:
                self._performance_rating_fields.append([key, _dict[key]])

        self._performance_rating_past_fields = {}
        _dict = self._your_past_org.groupby("2017 Performance Rating").groups
        for key in _dict:
            _key = key
            if key == "Unspecified":
                _key = "No Rating"
            self._performance_rating_past_fields.update({_key: _dict[key]})

        ## make talent cordinate fields.
        self._talent_cordinate_fields = []
        _dict = self._your_org.groupby("2020 Talent Coordinate").groups
        for key in _dict:
            _key = key
            if key == "Unspecified":
                _key = "No Coordinate"
            self._talent_cordinate_fields.append([_key, _dict[key]])

        self._talent_cordinate_past_fields = {}
        _dict = self._your_past_org.groupby("2017 Talent Coordinate").groups
        for key in _dict:
            _key = key
            if key == "Unspecified":
                _key = "No Coordinate"
            self._talent_cordinate_past_fields.update({_key: _dict[key]})

        ## make gender fields.
        self._gender_fields = []
        _dict = self._your_org.groupby("Gender").groups
        for key in _dict:
            self._gender_fields.append([key, _dict[key]])

        self._gender_past_fields = {}
        _dict = self._your_past_org.groupby("Gender").groups
        for key in _dict:
            self._gender_past_fields.update({key: _dict[key]})

        ## make Ethnicity fields.
        self._ethnicity_fields = []
        _dict = self._your_org.groupby("Ethnicity (US)").groups
        for key in _dict:
            if not key == "Non-US":
                self._ethnicity_fields.append([key, _dict[key]])
        try:
            self._ethnicity_fields.append(["Non-US", _dict["Non-US"]])
        except:
            pass

        self._ethnicity_past_fields = {}
        _dict = self._your_past_org.groupby("Ethnicity (US)").groups
        for key in _dict:
            self._ethnicity_past_fields.update({key: _dict[key]})

        ## make age group fields.
        self._age_fields = []
        _dict = self._your_org.groupby("Age Group").groups
        for key in _dict:
            self._age_fields.append([key, _dict[key]])

        self._age_past_fields = {}
        _dict = self._your_past_org.groupby("Age Group").groups
        for key in _dict:
            self._age_past_fields.update({key: _dict[key]})

        ## make country fields.
        self._country_fields = []
        _dict = self._your_org.groupby("Country").groups
        for key in _dict:
            self._country_fields.append([key, _dict[key]])

        self._country_past_fields = {}
        _dict = self._your_past_org.groupby("Country").groups
        for key in _dict:
            self._country_past_fields.update({key: _dict[key]})

        ## make kite fields.
        self._kite_fields = []
        _dict = self._your_org.groupby("Kite Employee Flag").groups
        for key in _dict:
            self._kite_fields.append([key, _dict[key]])

        self._kite_past_fields = {}
        _dict = self._your_past_org.groupby("Kite Employee Flag").groups
        for key in _dict:
            self._kite_past_fields.update({key: _dict[key]})

        ## make office type fields.
        self._office_fields = []
        _dict = self._your_org.groupby("Office Type").groups
        for key in _dict:
            self._office_fields.append([key, _dict[key]])

        self._office_past_fields = {}
        _dict = self._your_past_org.groupby("Office Type").groups
        for key in _dict:
            self._office_past_fields.update({key: _dict[key]})

        ## make Region fields.
        self._region_fields = []
        _dict = self._your_org.groupby("Location Level 2").groups
        for key in _dict:
            self._region_fields.append([key, _dict[key]])

        self._region_past_fields = {}
        _dict = self._your_past_org.groupby("Location Level 2").groups
        for key in _dict:
            self._region_past_fields.update({key: _dict[key]})

        ## make Department fields.
        self._department_fields = []
        _dict = self._your_org.groupby("Department Level 2").groups
        for key in _dict:
            self._department_fields.append([key, _dict[key]])

        self._department_past_fields = {}
        _dict = self._your_past_org.groupby("Department Level 2").groups
        for key in _dict:
            self._department_past_fields.update({key: _dict[key]})

        ## make Gender x Ethnicity (US) fields.
        self._gender_ethnicity_fields = []
        _dict = self._your_org.groupby(["Ethnicity (US)", "Gender"]).groups
        _new_dict = copy.deepcopy(_dict)
        for key in _dict:
            _temp = _new_dict.pop(key)
            if key[0] != "Non-US":
                _new_key = "{} {}".format(key[0], key[1])
                _new_dict.update({_new_key: _temp})
        for key in _new_dict:
            self._gender_ethnicity_fields.append([key, _new_dict[key]])

        self._gender_ethnicity_past_fields = {}
        _dict = self._your_past_org.groupby(["Ethnicity (US)", "Gender"]).groups
        _new_dict = copy.deepcopy(_dict)
        for key in _dict:
            _temp = _new_dict.pop(key)
            if key[0] != "Non-US":
                _new_key = "{} {}".format(key[0], key[1])
                _new_dict.update({_new_key: _temp})
        for key in _new_dict:
            self._gender_ethnicity_past_fields.update({key: _new_dict[key]})


        ## make affiliate fields.
        if self.use_affiliate:
            self._affiliate_fields = []
            self._affiliate_past_fields = {}
            temp_id = "GM Level {} ID"
            temp_name = "GM Level {} Org"
            key_list = []
            def get_org_id_by_name(name):
                return self.GM_levels[self.GM_levels["GM Org"] == name]["GM ID"].values[0]
            
            if not self.affiliate_second:
                _3_levels = self.GM_levels[self.GM_levels[temp_id.format(2)] == self._leader_id]
                _3_list = _3_levels[~(_3_levels["GM ID"] == self._leader_id)][temp_name.format(3)].drop_duplicates().values
                _3_list = sorted(_3_list)
                _3_last = []
                _3_complete = []
                for item in _3_list[:]:
                    if item.startswith("Kite"):
                        _3_last.append(item)
                    else:
                        _3_complete.append(item)
                for item in _3_last:
                    _3_complete.append(item)

                for _3_name in _3_complete:
                    key_list.append(_3_name)
                    _3_id = get_org_id_by_name(_3_name)
                    _4_levels = self.GM_levels[self.GM_levels[temp_id.format(3)] == _3_id]
                    _4_list = _4_levels[~(_4_levels["GM ID"] == _3_id)][temp_name.format(4)].values
                    _4_list = sorted(_4_list)
                    _4_last = []
                    _4_complete = []
                    for item in _4_list:
                        if item.startswith("Kite"):
                            _4_last.append(item)
                        else:
                            _4_complete.append(item)
                    for item in _4_last:
                        _4_complete.append(item)

                    for _4_name in _4_complete:
                        key_list.append("    " + _4_name)
                
                for key in key_list:
                    real_key = key.strip()
                    try:
                        _list = self._your_org[self._your_org[real_key] == 1].index
                        self._affiliate_fields.append([key, _list])
                    except:
                        pass
                    try:
                        _list = self._your_past_org[self._your_past_org[real_key] == 1].index
                        self._affiliate_past_fields.update({key: _list})
                    except:
                        pass
            else:
                _4_levels = self.GM_levels[self.GM_levels[temp_id.format(3)] == self._leader_id]
                _4_list = _4_levels[~(_4_levels["GM ID"] == self._leader_id)][temp_name.format(4)].values
                _4_list = sorted(_4_list)
                _4_last = []
                _4_complete = []
                for item in _4_list:
                    if item.startswith("Kite"):
                        _4_last.append(item)
                    else:
                        _4_complete.append(item)
                for item in _4_last:
                    _4_complete.append(item)

                for _4_name in _4_complete:
                    key_list.append(_4_name)

                for key in key_list:
                    try:
                        _list = self._your_org[self._your_org[key] == 1].index
                        self._affiliate_fields.append([key, _list])
                    except:
                        pass
                    try:
                        _list = self._your_past_org[self._your_past_org[key] == 1].index
                        self._affiliate_past_fields.update({key: _list})
                    except:
                        pass

        self.index_match = {
            "": ["Gilead Overall Delta", "Parent Group Delta", "Your Org Delta ({} to {})".format(self.current_year, self.past_year)],
            "Direct Reports (as of 2 Sept 2020)": sorted(self._get_names_from_field(self._direct_report_field)),
            "Grade Group": self._get_names_from_field(self._grade_group_fields),
            "Tenure Group": self._get_names_from_field(self._tenure_group_fields),
            "Office Type": self._get_names_from_field(self._office_fields),
            "Performance Rating": self._get_names_from_field(self._performance_rating_fields),
            "Talent Coordinate": self._get_names_from_field(self._talent_cordinate_fields),
            "Gender": self._get_names_from_field(self._gender_fields),
            "Ethnicity (US)": self._get_names_from_field(self._ethnicity_fields),
            "Gender x Ethnicity (US)": self._get_names_from_field(self._gender_ethnicity_fields),
            "Age Group": self._get_names_from_field(self._age_fields),
            "Function": self._get_names_from_field(self._department_fields),
            "Region": self._get_names_from_field(self._region_fields),
            "Country": self._get_names_from_field(self._country_fields),
            "Kite": self._get_names_from_field(self._kite_fields),
        }

        if self.logic == 2:
            self.index_match.update({"": ["Gilead Overall Delta", "Your Org Delta ({} to {})".format(self.current_year, self.past_year)]})
        elif self.logic == 1:
            self.index_match.update({"": ["Gilead Overall Delta"]})

        if self.GM:
            self.index_match.pop("Direct Reports (as of April 24, 2018)", "no")

        if self.use_affiliate:
            self.index_match.pop("Country", "no")
            _keys = list(self.index_match.keys())
            _keys.insert(_keys.index("Grade Group"), "Affiliate")
            _dict = self.index_match
            self.index_match = {}
            for key in _keys:
                if key == "Affiliate":
                    self.index_match[key] = self._get_names_from_field(self._affiliate_fields)
                else:
                    self.index_match[key] = _dict[key]

        ## make and initiate a dict to save all calculated data.
        self.precious_dict = {}
        for first_index in self.index_match:
            _ = {}
            for item in self.index_match[first_index]:
                _.update({item: {}})
            self.precious_dict.update({first_index: _})

    def _get_names_from_field(self, field_list):

        keys = [field[0] for field in field_list]
        return keys

    def _filterResource(self, id_list, category):

        # filter_item = self._item_pd[self._item_pd["Item ID"].isin(id_list)]["Unique Item Code"].tolist()
        # Hi May, the below bounded area is changed part.
        #########################################################################################################
        filter_item = []
        for item_id in id_list:
            filter_item.append(self._item_pd[self._item_pd["Item ID"] == item_id]["Unique Item Code"].values[0])
        #########################################################################################################
        _items  = ['ExternalReference']
        _items_past = ['ExternalReference']

        for item in filter_item:
            item_past = self._get_past_field_name_by_current_name(item)
            if item_past != None:
                _items.append(item)
                _items_past.append(item_past)
            else:
                pass

        if len(_items) > 1:
            self._item_list.append([0, category])
            for item in _items:
                if item != 'ExternalReference':
                    self._item_list.append([1, item])
        
        self._filtered_raw_data = self.raw_data_pd[_items]
        self._filtered_raw_past_data = self.raw_data_past_pd[_items_past]

    def _calculateEachRow(self, item):

        ## calculate Gilead overall delta %s
        _dict, c_lens, p_lens = self._calculateOverall(self._gilead_org, self._gilead_past_org, item)
        self.precious_dict[""]["Gilead Overall Delta"].update(_dict)
        self.first_row["current"].update({"Gilead Overall Delta": c_lens})
        self.first_row["past"].update({"Gilead Overall Delta": p_lens})

        ## calculate Parent Group delta %s
        if self.logic == 3:
            _dict, c_lens, p_lens = self._calculateOverall(self._parent_org, self._parent_past_org, item)
            self.precious_dict[""]["Parent Group Delta"].update(_dict)
            self.first_row["current"].update({"Parent Group Delta": c_lens})
            self.first_row["past"].update({"Parent Group Delta": p_lens})

        ## calculate Your Org Delta (2020 to 2018) %s
        if self.logic >= 2:
            _dict, c_lens, p_lens = self._calculateOverall(self._your_org, self._your_past_org, item)
            self.precious_dict[""]["Your Org Delta ({} to {})".format(self.current_year, self.past_year)].update(_dict)
            self.first_row["current"].update({"Your Org Delta ({} to {})".format(self.current_year, self.past_year): c_lens})
            self.first_row["past"].update({"Your Org Delta ({} to {})".format(self.current_year, self.past_year): p_lens})

        ## calculate Direct reports %s
        if not self.GM:
            self._calculateSubFields(self._direct_report_field, self._direct_report_past_field, "Direct Reports (as of 2 Sept 2020)", item)

        ## calcualte Grade Group %s
        self._calculateSubFields(self._grade_group_fields, self._grade_group_past_fields, "Grade Group", item)

        ## calcualte Tenure Group %s
        self._calculateSubFields(self._tenure_group_fields, self._tenure_group_past_fields, "Tenure Group", item)

        ## calculate Performance Rating %s
        self._calculateSubFields(self._performance_rating_fields, self._performance_rating_past_fields, "Performance Rating", item)

        ## calculate Talent Coordinate %s
        self._calculateSubFields(self._talent_cordinate_fields, self._talent_cordinate_past_fields, "Talent Coordinate", item)

        ## calculate Gender %s
        self._calculateSubFields(self._gender_fields, self._gender_past_fields, "Gender", item)

        ## calculate Ethnicity (US) %s
        self._calculateSubFields(self._ethnicity_fields, self._ethnicity_past_fields, "Ethnicity (US)", item)

        ## calculate Age Group %s
        self._calculateSubFields(self._age_fields, self._age_past_fields, "Age Group", item)

        ## calculate Country %s
        if not self.use_affiliate:
            self._calculateSubFields(self._country_fields, self._country_past_fields, "Country", item)

        ## calculate Kite %s
        self._calculateSubFields(self._kite_fields, self._kite_past_fields, "Kite", item)

        ## calculate office %s
        self._calculateSubFields(self._office_fields, self._office_past_fields, "Office Type", item)

        ## calculate regions %s
        self._calculateSubFields(self._region_fields, self._region_past_fields, "Region", item)

        ## calculate department %s
        self._calculateSubFields(self._department_fields, self._department_past_fields, "Function", item)

        ## calculate affiliate %s
        if self.use_affiliate:
            self._calculateSubFields(self._affiliate_fields, self._affiliate_past_fields, "Affiliate", item)

        ## calculate gender ethnicity %s
        self._calculateSubFields(self._gender_ethnicity_fields, self._gender_ethnicity_past_fields, "Gender x Ethnicity (US)", item)

    def _calculateOverall(self, current_f, past_f, item):
        
        ## calcualte overall fields.
        current_ids = current_f.iloc[:, 0]
        current_pd = self._filtered_raw_data[self._filtered_raw_data["ExternalReference"].isin(current_ids)].reset_index(drop=True)
        nums = len(current_pd.iloc[:, 0])
        current_dict = self._get_sum(current_pd, nums, item)

        past_ids = past_f.iloc[:, 0]
        past_pd = self._filtered_raw_past_data[self._filtered_raw_past_data["ExternalReference"].isin(past_ids)].reset_index(drop=True)
        nums = len(past_pd.iloc[:, 0])
        past_dict = self._get_sum(past_pd, nums, item)

        for key in current_dict:
            try:
                value = current_dict[key][0] - past_dict[self._get_past_field_name_by_current_name(key)][0]
                current_dict[key][0] = value
            except:
                current_dict[key][0] = "N/A"
        
        return current_dict, len(current_ids), len(past_ids)

    def _get_sum(self, data, nums, item):

        ## calculate the sum of favorable scores.
        _dict = {}

        for ind in range(len(data.columns)):
            _ = data.iloc[:, ind]
            is_empty_column = True
            lens = nums
            sub = 0
            count_valid = 0
            for __ in _:
                if __ != '':
                    is_empty_column = False
                    sub += __
                    count_valid += 1
                else:
                    lens -= 1
            if is_empty_column:
                _dict.update({data.columns.values[ind]: ["N/A", 1]})
            else:
                if count_valid >= 4:
                    _dict.update({data.columns.values[ind]: [sub / lens, 1]})
                else:
                    _dict.update({data.columns.values[ind]: ["N/A", 1]})
        
        if len(data.columns) > 0:
            _dict.update({item: ["N/A", 0]})

        return _dict

    def _calculateSubFields(self, dataframe, pastframe, column_name, item):

        ## calculate the sum of all sub fields except overall fields.
        for field in dataframe:
            try:
                dictionary = self.precious_dict[column_name]
                c_list = field[1]
                c_ids = self._your_org.iloc[c_list, 0].tolist()
                c_nums = len(c_ids)
                c_pd = self._filtered_raw_data[self._filtered_raw_data["ExternalReference"].isin(c_ids)].reset_index(drop=True)
                c_dict = self._get_sum(c_pd.iloc[:, 1:], c_nums, item)

                p_list = pastframe[field[0]]
                p_ids = self._your_past_org.iloc[p_list, 0].tolist()
                p_nums = len(p_ids)
                p_pd = self._filtered_raw_past_data[self._filtered_raw_past_data["ExternalReference"].isin(p_ids)].reset_index(drop=True)
                p_dict = self._get_sum(p_pd.iloc[:, 1:], p_nums, item)

                for key in c_dict:
                    try:
                        value = c_dict[key][0] - p_dict[self._get_past_field_name_by_current_name(key)][0]
                        c_dict[key][0] = value
                    except:
                        c_dict[key][0] = "N/A"
                
                dictionary[field[0]].update(c_dict)
                self.first_row["current"].update({column_name + field[0]: c_nums})
                self.first_row["past"].update({column_name + field[0]: p_nums})

            except:
                pass

    def _get_color(self, delta):
        if abs(delta) > 25:
            delta = delta // abs(delta) * 25
        series = self.heatmap_color_pd[self.heatmap_color_pd["Delta"] == delta]
        return str(hex(series["R"].values[0]))[2:] + str(hex(series["G"].values[0]))[2:] + str(hex(series["B"].values[0]))[2:]


class SSM:

    def __init__(self, **args):
        ## initialize the object by specifying input and output files.
        self.image_src = args['image']

        self.raw_data_file = args['raw_data']
        self.raw_data_past_file = args['raw_data_past']
        self.item_code_file = args['item_code']
        self.demographics_file = args['demographics']
        self.demographics_past_file = args['demographics_past']
        self.heatmap_color_file = args['heatmap_color']
        self.benchmark_file = args['benchmark']
        self.leader_file = args['leader']
        self.output_source = args['output_folder']
        self.input_source = args['input_folder']
        self.how2use_file = args['how to use']

        self.custom_text = args['custom text']

        self.current_year = self.demographics_file[:4]
        self.past_year = self.demographics_past_file[:4]

    def readAllFiles(self):
        ## read files and save it in object data.
        self.origin_raw_data_pd = pd.read_excel(self.input_source + "/" + self.raw_data_file, engine="openpyxl")
        self.origin_raw_data_past_pd = pd.read_excel(self.input_source + "/" + self.raw_data_past_file, engine="openpyxl")
        self.item_code_pd = pd.read_excel(self.input_source + "/" + self.item_code_file, engine="openpyxl", sheet_name="ItemCodeSTAR")
        self.origin_category_pd =  pd.read_excel(self.input_source + "/" + self.item_code_file, engine="openpyxl", sheet_name="CurrentCategorySTAR")
        self.origin_demographics_pd = pd.read_excel(self.input_source + "/" + self.demographics_file, engine="openpyxl")
        self.demographics_past_pd = pd.read_excel(self.input_source + "/" + self.demographics_past_file, engine="openpyxl")
        self.heatmap_color_pd = pd.read_excel(self.input_source + "/" + self.heatmap_color_file, engine="openpyxl")
        self.benchmark_pd = pd.read_excel(self.input_source + "/" + self.benchmark_file, engine="openpyxl")
        self.leaders = pd.read_excel(self.input_source + "/" + self.leader_file, engine="openpyxl", sheet_name="Leader")
        self.GMs = pd.read_excel(self.input_source + "/" + self.leader_file, engine="openpyxl", sheet_name="GM")
        self.site_leads = pd.read_excel(self.input_source + "/" + self.leader_file, engine="openpyxl", sheet_name="Site Leader")
        self.how2use_pd = pd.read_excel(self.input_source + "/" + self.how2use_file, engine="openpyxl", sheet_name="Score Summary How to Use")

    def setLeader(self, id, GM=False, site_lead=False):
        self._leader_id = id
        self.GM = GM
        self.site_lead = site_lead

        # self._leader_id = 112372

    def calculateValues(self):
        ## do some process referred to individual leader ID.
        self._preProcess()
        self._prepareColumnsForID()

        item_list = []
        item_dict = {}
        category_list = []

        ## fill the above variables.
        for key in self.order_category:
            category_list.append(key)
            item_list.append([0, key])
            temp_list = []
            for item in self.category_pd.iloc[self._group_dict[key], 0]:
                item_list.append([1, item])
                temp_list.append(item)
            item_dict.update({key:temp_list})

        self._item_list = item_list

        ## do main calculation process by iterating over the category list.
        for category in tqdm(category_list, desc="iterating over the list of category"):
            sub_id_list = item_dict[category]

            ## filter the source data frame to have columns belonged to each category.
            self._filterResource(sub_id_list)

            ## based on the filtered source, calculate the values for each row.
            self._calculateEachRow(category)
        

        self._makeBenchColumn()

    def makeReport(self):

        ## prepare image.
        img  = Image(self.input_source + "/" + self.image_src)
        img.height = 90
        img.width = 110

        ## calculate total rows that will be placed in our output file.
        total_rows = 7 + len(self._item_list)

        ## make a workbook and sheet.
        self.book = openpyxl.Workbook()
        sheet = self.book.active
        sheet.title = "Score Summary"

        ## set styles like font, color, direction, border...
        ft = Font(name="Arial", size=10)
        ft_bold = Font(name="Arial", size=10, bold=True)
        ft_size_bold = Font(name="Arial", size=11, bold=True)
        ft_tiny = Font(name="Arial", size=8)
        side = Side(style='thin', color="CCCCCC")
        category_border = Border(
                     top=side,
                     bottom=side)
        last_border = Border(bottom=side)
        white_back = PatternFill("solid", fgColor="FFFFFF")
        center_alignment = Alignment(horizontal='center', vertical="center")
        right_alignment = Alignment(horizontal='right', vertical='center')

        ## set empty cells white.
        for row in range(1, total_rows + 40 + 1):
            sheet.row_dimensions[row].height = 13
            for col in range(1, 80):
                cell = sheet.cell(row=row, column=col)
                cell.fill = white_back
                cell.alignment = center_alignment
                cell.font = ft
        
        for col in range(1, 80):
            sheet.column_dimensions[ce.get_column_letter(col)].width = 6.5

        ## set width and height of columns.
        # sheet.column_dimensions[ce.get_column_letter(1)].width = 40
        sheet.column_dimensions[ce.get_column_letter(1)].width = 6
        sheet.column_dimensions[ce.get_column_letter(2)].width = 4
        sheet.column_dimensions[ce.get_column_letter(3)].width = 26
        sheet.column_dimensions[ce.get_column_letter(4)].width = 4

        # sheet.column_dimensions[ce.get_column_letter(2)].width = 40
        sheet.column_dimensions[ce.get_column_letter(5)].width = 20
        sheet.column_dimensions[ce.get_column_letter(6)].width = 4
        sheet.column_dimensions[ce.get_column_letter(7)].width = 16

        sheet.column_dimensions[ce.get_column_letter(8)].width = 1

        ## insert image.
        sheet.row_dimensions[2].height = 20
        size = XDRPositiveSize2D(p2e(80), p2e(60))
        marker = AnchorMarker(col=9, colOff=p2e(0), row=0, rowOff=p2e(10))
        img.anchor = OneCellAnchor(_from=marker, ext=size)

        sheet.add_image(img)

        ## fill fifth row.
        __value = "Your Org"
        if self.GM:
            __value = self.GM
        elif self.site_lead:
            __value = self.site_lead
        if self.logic == 1:
            __value = "Gilead Overall"
        sheet.merge_cells(start_row=5, end_row=5, start_column=5, end_column=7)
        sheet.merge_cells(start_row=6, end_row=6, start_column=5, end_column=7)
        cell = sheet.cell(row=5, column=4 + 1)
        cell.value = __value
        cell.alignment = center_alignment
        cell.font = ft_size_bold

        sheet.merge_cells(start_row=5, start_column=9, end_row=5, end_column=11)
        cell = sheet.cell(row=5, column=9)
        cell.value = "Deltas*"
        cell.font = ft_size_bold

        ## fill sixth row.
        _string = "n = " + f"{self._participated:,d}" + ' / ' + f"{self._invited:,d}" + " ({}% participation)".format(round(self._participated / self._invited * 100))
        sheet.cell(row=6, column=4 + 1).value = _string
        sheet.cell(row=6, column=7 + 1 + 1).value = self.past_year
        sheet.cell(row=6, column=7 + 1 + 2).value = "Ext"

        _value = "Gilead"
        if self.logic == 1:
            _value = "Kite"
        sheet.cell(row=6, column=7 + 1 + 3).value = _value

        ## fill all data.
        for index, (criteria, item) in enumerate(tqdm(self._item_list, desc="making contents...")):
            ## fill the first column.
            row_number = index + 8
            cell = sheet.cell(row=row_number, column=1)
            cell.alignment = right_alignment
            cell.value = item
            if criteria == 1:
                cell.value = self._item_pd[self._item_pd["Item ID"] == item]["Short Text [2020 onward]"].values[0]
                item = self._item_pd[self._item_pd["Item ID"] == item]["Unique Item Code"].values[0]

            cell.font = ft
            if index == len(self._item_list) - 1:
                cell.border = last_border
                sheet.cell(row=row_number, column=4 + 1).border = last_border

            if criteria == 0:
                cell.border = category_border
                cell.font = ft_bold

                sheet.cell(row=row_number, column=4 + 1).border = category_border
            
            ## fill the third group columns.
            sub_index = 0
            for key in self.left_dict:
                cell = sheet.cell(row=row_number, column=7 + 1 + 1 + sub_index)
                try:
                    value = self.left_dict[key][item][0]
                except:
                    value = "N/A"
                cell.value = value if value != "N/A" else "-"
                cell.number_format = numbers.FORMAT_PERCENTAGE
                cell.font = ft
                cell.alignment = center_alignment

                try:
                    if value >= 0:
                        cell.number_format = "+0%"
                    elif value < 0 and value > -0.005:
                        cell.number_format = "-0%"
                except:
                    pass

                try:
                    cell.fill = PatternFill("solid", fgColor=self._get_color(round(value * 100)))
                except:
                    ## this skip the case of N/A
                    pass

                if index == len(self._item_list) - 1:
                    cell.border = last_border
                if criteria == 0:
                    cell.border = category_border
                sub_index += 1
                
            ## fill the last group columns.
            sub_index = 0
            for key in self.right_dict:
                cell = sheet.cell(row=row_number, column=37 + sub_index)
                value = self.right_dict[key][item][0]
                cell.value = value if value != "N/A" else "-"
                cell.number_format = numbers.FORMAT_PERCENTAGE
                cell.font = ft
                cell.alignment = center_alignment

                if index == len(self._item_list) - 1:
                    cell.border = last_border
                if criteria == 0:
                    cell.border = category_border
                sub_index += 1

        ## add caption field.
        sheet.merge_cells(start_row=2, end_row=3, start_column=1, end_column=7)
        cell = sheet.cell(row=2, column=1)
        cell.value = "Global Employee Survey Results"
        cell.font = Font(name="Arial Black", size=24, color="D9D9D9", bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

        ## merge cells(first step).
        for index in range(total_rows - 7):
            _row = 8 + index
            sheet.merge_cells(start_column=1, end_column=4, start_row=_row, end_row=_row)
            sheet.merge_cells(start_column=5, end_column=7, start_row=_row, end_row=_row)


        ## add chart bar.
        _data = Reference(sheet, min_col=37, min_row=8, max_col=39, max_row=total_rows)
        chart = BarChart()
        chart.add_data(_data)
        chart.height = 16.9
        chart.width = 7.6
        chart.type = "bar"
        chart.gapWidth = 50.0
        chart.grouping = "percentStacked"
        chart.overlap = 100
        chart.legend = None
        chart.y_axis.majorGridlines = None
        chart.y_axis.delete = True
        chart.x_axis.scaling.orientation = "maxMin"
        chart.x_axis.delete = True
        # chart.y_axis.font = Font(name="aakar", size=15)

        s = chart.series[0]
        s.graphicalProperties.line.solidFill = "7f9ba7"
        s.graphicalProperties.solidFill = "7f9ba7"
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.showCatName = False
        s.dLbls.showLegendkey = False
        s.dLbls.numFmt = "0%"
        # s.font = Font(name="aakar", size=15)

        s = chart.series[1]
        s.graphicalProperties.line.solidFill = "d8dada"
        s.graphicalProperties.solidFill = "d8dada"

        s = chart.series[2]
        s.graphicalProperties.line.solidFill = "c2bfb5"
        s.graphicalProperties.solidFill = "c2bfb5"  

        sheet.add_chart(chart, "E7")

        ## insert custom text
        __value = "your org"
        if self.GM:
            __value = self.GM
            __value = __value[:-3] + "org"
        elif self.site_lead:
            __value = self.site_lead
        if self.logic == 1:
            __value = "Gilead Overall"
        _texts = self.custom_text.format(__value, __value).split("\n")
        if self.logic == 1:
            _texts[-2] = "Kite compares Kite scores to rest of Gilead (negative indicates Kite less favorable)"
        for index, _text in enumerate(_texts):
            cell = sheet.cell(row=46 + index, column=1)
            cell.value = _text
            cell.alignment = Alignment(shrinkToFit=False)
            cell.font = ft_tiny

        # insert legend part.
        cell = sheet.cell(row=total_rows + 2, column=3)
        cell.value = "Favorable (Agree / Strongly Agree)"
        cell.font = ft_tiny
        cell.alignment = Alignment(horizontal="left", vertical="center", shrinkToFit=False)

        cell = sheet.cell(row=total_rows + 2, column=5)
        cell.value = "Neither Agree Nor Disagree"
        cell.font = ft_tiny
        cell.alignment = Alignment(horizontal="left", vertical="center", shrinkToFit=False)

        cell = sheet.cell(row=total_rows + 2, column=7)
        cell.value = "Unfavorable (Disagree / Strongly Disagree)"
        cell.font = ft_tiny
        cell.alignment = Alignment(horizontal="left", vertical="center", shrinkToFit=False)

        sheet.cell(row=total_rows + 2, column=2).fill = PatternFill("solid", fgColor="7f9ba7")
        sheet.cell(row=total_rows + 2, column=4).fill = PatternFill("solid", fgColor="d8dada")
        sheet.cell(row=total_rows + 2, column=6).fill = PatternFill("solid", fgColor="c2bfb5")

        ## insert 'how to use' sheet
        how_sheet = self.book.create_sheet("How to Use")
        how_font = Font(name="Arial", size=10)
        cell = how_sheet.cell(column=1, row=1)
        cell.value = self.how2use_pd.columns.values[0]
        cell.alignment = Alignment(wrapText=True, shrinkToFit=False, indent=0)
        cell.font = how_font
        how_sheet.row_dimensions[1].height = int(len(self.how2use_pd.columns.values[0]) / 120 * (40 / 3))
        

        for index in range(len(self.how2use_pd.index)):
            _ = 0
            for col_index in range(len(self.how2use_pd.columns)):
                cell = how_sheet.cell(column=1 + col_index, row=2 + index)

                content = self.how2use_pd.iloc[index, col_index]
                if _ < len(content):
                    _ = len(content)
                cell.value = content
                cell.alignment = Alignment(wrapText=True, shrinkToFit=False)
                cell.font = how_font
                how_sheet.column_dimensions[ce.get_column_letter(col_index + 1)].width = 135

            how_sheet.row_dimensions[index + 2].height = int(_ / 110 * (40 / 3))



        for row in range(1, 41 + len(self.how2use_pd.index)):
            for col in range(1, 41 + len(self.how2use_pd.columns)):
                how_sheet.cell(row=row, column=col).fill = white_back

    def writeOutput(self):
        ## specify the path of output file
        if self.GM:
            self.file_name = self.GM[:-4]
            self.output_path = "/" + self.file_name
        if self.site_lead:
            self.file_name = self.site_lead.replace(" / ", " ")
            self.output_path = "/" + self.file_name
        path = self.output_source + self.output_path + "/" + "2021-04 Global Employee Survey - " + self.file_name + " - Score Summary.xlsx"

        ## if the output file already exists, remove it.
        if os.path.exists(path):
            os.remove(path)

        ## make a folder to involve the output file.
        os.makedirs(self.output_source + self.output_path, exist_ok=True)

        ## and write the output file.
        self.book.save(path)

    def _preProcess(self):
        ## make a item group and filter the able source.
        self._item_pd = self.item_code_pd[self.item_code_pd["Type ID"] == "T01"]

        self._benchmark_item_pd = self._item_pd[self._item_pd["External Benchmark"] == "e"]
        self._item_pd = self._item_pd[self._item_pd["External Benchmark"] == "i"]

        self._rest_item_pd = self._item_pd[~self._item_pd["Unique Item Code"].isin(self.origin_raw_data_pd.columns.values)].reset_index(drop=True)
        self._item_pd = self._item_pd[self._item_pd["Unique Item Code"].isin(self.origin_raw_data_pd.columns.values)].reset_index(drop=True)
        self._rest_item_pd = self._rest_item_pd[self._rest_item_pd["Unique Item Code"].isin(self.origin_raw_data_past_pd.columns.values)].reset_index(drop=True)
        self._both_item_pd = self._item_pd[self._item_pd["Unique Item Code"].isin(self.origin_raw_data_past_pd.columns.values)].reset_index(drop=True)

        self.category_pd = self.origin_category_pd[self.origin_category_pd["Item ID in 2020 Survey"].isin(self._item_pd["Item ID"].tolist())]
        self.category_pd = self.category_pd.drop_duplicates(subset=["Item ID in 2020 Survey"]).reset_index(drop=True)
        self.order_category = self.category_pd.drop_duplicates(subset=["2020 Category"]).loc[:, "2020 Category"].tolist()

        self._group_dict = self.category_pd.groupby(["2020 Category"]).groups

        self.raw_data_pd = self.origin_raw_data_pd.iloc[2:].reset_index(drop=True)

        ## filter the able source from history file.
        self.raw_data_past_pd = self.origin_raw_data_past_pd.iloc[2:].reset_index(drop=True)

        ## Convert numeric values into favorable, neutral, or not.
        self.raw_data_n_pd = self.raw_data_pd.copy()
        self.raw_data_uf_pd = self.raw_data_pd.copy()

        for field in self._item_pd["Unique Item Code"]:
            new_list = [[], [], []]
            for item in self.raw_data_pd[field].tolist():
                if item == 4 or item == 5:
                    new_list[0].append(1)
                    new_list[1].append(0)
                    new_list[2].append(0)

                elif item == 3:
                    new_list[0].append(0)
                    new_list[1].append(1)
                    new_list[2].append(0)

                elif item == 1 or item == 2:
                    new_list[0].append(0)
                    new_list[1].append(0)
                    new_list[2].append(1)

                else:
                    new_list[0].append('')
                    new_list[1].append('')
                    new_list[2].append('')

            self.raw_data_pd[field] = new_list[0]
            self.raw_data_n_pd[field] = new_list[1]
            self.raw_data_uf_pd[field] = new_list[2]

            ## do the same process about history data.
            try:
                field = self._get_past_field_name_by_current_name(field)
                new_list_past = []
                for item in self.raw_data_past_pd[field].tolist():
                    if item < 4 and item > 0:
                        new_list_past.append(0)
                    elif item >= 4 and item <= 5:
                        new_list_past.append(1)
                    else:
                        new_list_past.append('')
                self.raw_data_past_pd[field] = new_list_past
            except:
                pass

        ## new feature -> process A/B pair.
        pairs = self._item_pd[self._item_pd["AB Code"].isin(["A", "B"])].reset_index(drop=True)
        pairs_dict = pairs.groupby(["Item ID"]).groups

        for key in pairs_dict:

            _list = pairs_dict[key]
            pairs_row = pairs.iloc[_list, :]
            item_list = pairs_row["Unique Item Code"].tolist()
            text_list = pairs_row["Short Text [2020 onward]"].tolist()
            new_text = "/".join(text_list)
            self._item_pd = self._item_pd[~(self._item_pd["Unique Item Code"] == item_list[1])].reset_index(drop=True)
            _index = self._item_pd[self._item_pd["Unique Item Code"] == item_list[0]].index
            self._item_pd.loc[_index, "Short Text [2020 onward]"] = new_text

            column_pair = self.raw_data_pd[item_list]
            column_n_pair = self.raw_data_n_pd[item_list]
            column_uf_pair = self.raw_data_uf_pd[item_list]

            _list = []
            for index in range(len(column_pair.index)):
                _row = column_pair.iloc[index, :].tolist()
                for val in _row:
                    if type(val) == type(0):
                        _list.append(val)
                        break
                else:
                    _list.append('')

            _pd = pd.DataFrame(_list, columns=[item_list[0]])
            self.raw_data_pd = self.raw_data_pd.drop(columns=[item_list[1]])
            self.raw_data_pd[item_list[0]] = _pd

            _list = []
            for index in range(len(column_n_pair.index)):
                _row = column_n_pair.iloc[index, :].tolist()
                for val in _row:
                    if type(val) == type(0):
                        _list.append(val)
                        break
                else:
                    _list.append('')

            _pd = pd.DataFrame(_list, columns=[item_list[0]])
            self.raw_data_n_pd = self.raw_data_n_pd.drop(columns=[item_list[1]])
            self.raw_data_n_pd[item_list[0]] = _pd

            _list = []
            for index in range(len(column_uf_pair.index)):
                _row = column_uf_pair.iloc[index, :].tolist()
                for val in _row:
                    if type(val) == type(0):
                        _list.append(val)
                        break
                else:
                    _list.append('')

            _pd = pd.DataFrame(_list, columns=[item_list[0]])
            self.raw_data_uf_pd = self.raw_data_uf_pd.drop(columns=[item_list[1]])
            self.raw_data_uf_pd[item_list[0]] = _pd

        self.demographics_pd = self.origin_demographics_pd

        ## convert the demographics data to include only answered entries.
        self._invited_demographics_data = self.demographics_pd[self.demographics_pd["Invitee Flag"] == 1].reset_index(drop=True)
        self._answered_demographics_data = self._invited_demographics_data[self._invited_demographics_data.loc[:, "Worker ID"].isin(self.raw_data_pd['ExternalReference'].tolist())].reset_index(drop=True)
        self._gilead_org = self._answered_demographics_data

        ## do the same process about history data.
        self._invited_demographics_past_data = self.demographics_past_pd[self.demographics_past_pd["Invitee Flag"] == 1].reset_index(drop=True)
        self._answered_demographics_past_data = self._invited_demographics_past_data[self._invited_demographics_past_data.loc[:, "Worker ID"].isin(self.raw_data_past_pd['ExternalReference'].tolist())].reset_index(drop=True)

        if self.GM:
            self._gm_demographics_data = self._answered_demographics_data[self._answered_demographics_data[self.GM] == 1].reset_index(drop=True)
            self._gm_demographics_past_data = self._answered_demographics_past_data[self._answered_demographics_past_data[self.GM] == 1].reset_index(drop=True)

        if self.site_lead:
            self._site_demographics_data = self._answered_demographics_data[self._answered_demographics_data[self.site_lead] == 1].reset_index(drop=True)
            self._site_demographics_past_data = self._answered_demographics_past_data[self._answered_demographics_past_data[self.site_lead] == 1].reset_index(drop=True)      

        self.right_dict = {'f':{}, 'n':{}, 'uf':{}}
        self.left_dict = {'d':{}, 'e':{}, 'f':{}}
        self.benchmark_dict = {}
        self.past_dict = {}

    def _get_past_field_name_by_current_name(self, field_name):
        _item_id = self._item_pd[self._item_pd["Unique Item Code"] == field_name]["Item ID"].values[0]
        try:
            new_field_name = self._rest_item_pd[self._rest_item_pd["Item ID"] == _item_id]["Unique Item Code"].values[0]
        except:
            new_field_name = None

        if new_field_name == None:
            try:
                new_field_name = self._both_item_pd[self._both_item_pd["Unique Item Code"] == field_name]["Unique Item Code"].values[0]
            except:
                new_field_name = None

        return new_field_name

    def _get_benchmark_field_name_by_current_name(self, field_name):
        _item_id = self._item_pd[self._item_pd["Unique Item Code"] == field_name]["Item ID"].values[0]
        try:
            field_name = self._benchmark_item_pd[self._benchmark_item_pd["Item ID"] == _item_id]["Unique Item Code"].values[0]
        except:
            field_name = None
        
        return field_name

    def _prepareColumnsForID(self):
        ## find the supervisor level of the given leader id.
        self.first_row = {"current": {}, "past": {}}
        _temp = "Supervisor Level {} ID"

        leader_entry = self.origin_demographics_pd[self.origin_demographics_pd.loc[:, "Worker ID"] == self._leader_id]

        if self._leader_id == 999999:
            self.logic = 1
            leader_level = 1
        else:
            for index, level in enumerate(leader_entry.loc[:, _temp.format(2) : _temp.format(10)]):
                if (leader_entry[level] == self._leader_id).tolist()[0]:
                    leader_level = index + 2
                    break

        self.output_path = "/" + leader_entry["Worker Name"].values[0]
        self.file_name = leader_entry["Worker Last Name"].values[0]

        supervisor_level = leader_level - 1

        if leader_level >= 3:
            self.logic = 3
            self._supervisor_id = leader_entry.loc[:, _temp.format(supervisor_level)].values[0]
            _supervisor_entry = self.origin_demographics_pd[self.origin_demographics_pd.loc[:, "Worker ID"] == self._supervisor_id]
            self._supervisor_last_name = _supervisor_entry["Worker Last Name"].values[0]

            ## get Parent group.
            # self._parent_org = self._answered_demographics_data[self._answered_demographics_data.loc[:, _temp.format(supervisor_level)] == self._supervisor_id].reset_index(drop=True)
            # self._parent_past_org = self._answered_demographics_past_data[self._answered_demographics_past_data.loc[:, _temp.format(supervisor_level)] == self._supervisor_id].reset_index(drop=True)

        elif leader_level == 2:
            self.logic = 2

        ## get Your Org data
        if self.logic == 1:
            self._your_org = self._answered_demographics_data
            self._invited = len(self._invited_demographics_data.index)
            self._invited_past = len(self._invited_demographics_past_data.index)

            self._kite = self._answered_demographics_data[self._answered_demographics_data["Kite Employee Flag"] == "Kite"].reset_index(drop=True)
            self._no_kite = self._answered_demographics_data[self._answered_demographics_data["Kite Employee Flag"] == "Gilead (No Kite)"].reset_index(drop=True)
        else:
            if self.GM:
                self._your_org = self._gm_demographics_data
                self._invited = len(self._invited_demographics_data[self._invited_demographics_data[self.GM] == 1].index)
                self._invited_past = len(self._invited_demographics_past_data[self._invited_demographics_past_data[self.GM] == 1].index)

            elif self.site_lead:
                self._your_org = self._site_demographics_data
                self._invited = len(self._invited_demographics_data[self._invited_demographics_data[self.site_lead] == 1].index)

            else:
                self._your_org = self._answered_demographics_data[self._answered_demographics_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True)
                self._invited = len(self._invited_demographics_data[self._invited_demographics_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True).iloc[:, 0])
                self._invited_past = len(self._invited_demographics_past_data[self._invited_demographics_past_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True).iloc[:, 0])

        ## get your history group
        if self.GM:
            self._your_past_org = self._gm_demographics_past_data
        elif self.site_lead:
            self._your_past_org = self._site_demographics_past_data
        else:
            if self.logic == 1:
                self._your_past_org = self._answered_demographics_past_data
            else:
                self._your_past_org = self._answered_demographics_past_data[self._answered_demographics_past_data.loc[:, _temp.format(leader_level)] == self._leader_id].reset_index(drop=True)

        ## calculate nums of participated
        self._participated = len(self._your_org.loc[:, "Worker ID"])
        self._participated_past = len(self._your_past_org.loc[:, "Worker ID"])

    def _filterResource(self, id_list):

        filter_item = self._item_pd[self._item_pd["Item ID"].isin(id_list)]["Unique Item Code"].tolist()

        filter_item.insert(0, 'ExternalReference')
        self._filtered_raw_data = self.raw_data_pd[filter_item]
        self._filtered_raw_n_data = self.raw_data_n_pd[filter_item]
        self._filtered_raw_uf_data = self.raw_data_uf_pd[filter_item]

        ## about history data.
        self._helper_past = False
        filter_past_item = []
        for index, item in enumerate(filter_item):
            if index == 0:
                filter_past_item.append(item)
            else:
                new_item = self._get_past_field_name_by_current_name(item)
                if new_item != None:
                    filter_past_item.append(new_item)
                else:
                    self._helper_past = True
        
        self._filtered_raw_past_data = self.raw_data_past_pd[filter_past_item]

        ## about benchmark data.
        self._helper_benchmark = False
        filter_benchmark_item = []
        for index, item in enumerate(filter_item):
            if index == 0:
                filter_benchmark_item.append(item)
            else:
                new_item = self._get_benchmark_field_name_by_current_name(item)
                if new_item:
                    filter_benchmark_item.append(new_item)
                else:
                    self._helper_benchmark = True
        
        self._filtered_benchmark_value = self.benchmark_pd[self.benchmark_pd["Unique Item Code"].isin(filter_benchmark_item)]

    def _calculateEachRow(self, item):

        ## calculate the right columns.
        _dict, lens = self._calculateOverall(self._your_org, item)
        self.right_dict['f'].update(_dict)

        _dict, lens = self._calculateOverall(self._your_org, item, f=1)
        self.right_dict['n'].update(_dict)

        _dict, lens = self._calculateOverall(self._your_org, item, f=2)
        self.right_dict['uf'].update(_dict)

        ## calculate D %s
        if len(self._filtered_raw_past_data.columns) > 1:
            _dict, lens = self._calculateOverall(self._your_past_org, item, history=True)
            # for key in _dict:
            #     try:
            #         _dict[key][0] = self.right_dict['f'][key][0] - _dict[key][0]
            #     except:
            #         _dict[key][0] = 'N/A'
            # self.left_dict['d'].update(_dict)
            self.past_dict.update(_dict)
        
        ## calculate E %s
        if len(self._filtered_benchmark_value.columns) > 1:
            _dict = {}
            for index, row in self._filtered_benchmark_value.iterrows():
                value = row["External - CAmp Biotechnology & Medical Devices 2019"]
                _dict.update({row["Unique Item Code"]: [value, 1]})
            _dict.update({item: ["N/A", 0]})
            self.benchmark_dict.update(_dict)
            

        ## calculate F %s
        if self.logic == 1:
            _dict, lens = self._calculateOverall(self._kite, item)
            __dict, lens = self._calculateOverall(self._no_kite, item)

            for key in _dict:
                try:
                    _dict[key][0] = _dict[key][0] - __dict[key][0]
                except:
                    _dict[key][0] = "N/A"
        else:
            _dict, lens = self._calculateOverall(self._gilead_org, item)
            for key in _dict:
                try:
                    _dict[key][0] = self.right_dict['f'][key][0] - _dict[key][0]
                except:
                    _dict[key][0] = 'N/A'
        self.left_dict['f'].update(_dict)            


    def _calculateOverall(self, dataframe, item, history=False, f=0):

        ## calcualte overall fields.
        ids = dataframe.iloc[:, 0]
        if history:
            working_pd = self._filtered_raw_past_data[self._filtered_raw_past_data["ExternalReference"].isin(ids)].reset_index(drop=True)
        else:
            if f == 0:
                working_pd = self._filtered_raw_data[self._filtered_raw_data["ExternalReference"].isin(ids)].reset_index(drop=True)
            elif f == 1:
                working_pd = self._filtered_raw_n_data[self._filtered_raw_n_data["ExternalReference"].isin(ids)].reset_index(drop=True)
            elif f == 2:
                working_pd = self._filtered_raw_uf_data[self._filtered_raw_uf_data["ExternalReference"].isin(ids)].reset_index(drop=True)

        nums = len(working_pd.iloc[:, 0])
        return self._get_sum(working_pd.iloc[:, 1:], nums, item, history), len(ids)

    def _get_sum(self, data, nums, item, implicity=False):

        ## calculate the sum of favorable scores.
        _dict = {}
        determine_parent_na = False

        for ind in range(len(data.columns)):
            _ = data.iloc[:, ind]
            is_empty_column = True
            lens = nums
            sub = 0
            count_valid = 0
            for __ in _:
                if __ != '':
                    is_empty_column = False
                    sub += __
                    count_valid += 1
                else:
                    lens -= 1
            if is_empty_column:
                _dict.update({data.columns.values[ind]: ["N/A", 1]})
                determine_parent_na = True
            else:
                if count_valid >= 4:
                    _dict.update({data.columns.values[ind]: [sub / lens, 1]})
                else:
                    _dict.update({data.columns.values[ind]: ["N/A", 1]})
                    determine_parent_na = True
        
        if determine_parent_na:
            _dict.update({item: ["N/A", 0]})
        else:
            total_lens = nums
            total = 0
            cols = len(data.columns)
            for ind in range(nums):
                _series = data.iloc[ind, :]
                sub_sum = 0
                _is_nan = False
                for val in _series:
                    if val != '':
                        sub_sum += val
                    else:
                        _is_nan = True
                        total_lens -= 1
                        break

                if not _is_nan:
                    total += sub_sum / cols

            if total_lens < 4:
                _dict.update({item: ["N/A", 0]})
            else:
                _dict.update({item: [total / total_lens, 0]})

        if implicity:
                _dict.update({item: ["N/A", 0]})
        return _dict

    def _makeBenchColumn(self):
        for criteria, item in self._item_list:
            origin_item = item
            if criteria == 1:
                origin_item = self._item_pd[self._item_pd["Item ID"] == item]["Unique Item Code"].values[0]
                item = self._get_benchmark_field_name_by_current_name(origin_item)
            _ = ["N/A", 1 if criteria else 0]
            try:
                _ = self.benchmark_dict[item]
                if _[0] != "N/A":
                    _[0] = self.right_dict['f'][origin_item][0] - _[0]
            except:
                _ = ["N/A", 1 if criteria else 0]
            self.left_dict['e'].update({origin_item: _})

        for criteria, item in self._item_list:
            origin_item = item
            if criteria == 1:
                origin_item = self._item_pd[self._item_pd["Item ID"] == item]["Unique Item Code"].values[0]
                item = self._get_past_field_name_by_current_name(origin_item)
            _ = ["N/A", 1 if criteria else 0]
            try:
                _ = self.past_dict[item]
                if _[0] != "N/A":
                    _[0] = self.right_dict['f'][origin_item][0] - _[0]
            except:
                _ = ["N/A", 1 if criteria else 0]
            self.left_dict['d'].update({origin_item: _})


    def _get_color(self, delta):
        if abs(delta) > 25:
            delta = delta // abs(delta) * 25
        series = self.heatmap_color_pd[self.heatmap_color_pd["Delta"] == delta]
        return str(hex(series["R"].values[0]))[2:] + str(hex(series["G"].values[0]))[2:] + str(hex(series["B"].values[0]))[2:]

if __name__ == "__main__":

    ## Create a object.
    init_data = {
        'leader': "List of Leaders and GMs 2021-02-28.xlsx",
        'raw_data': "2020 Employee Survey Responses Sample 2021-02-05.xlsx",
        'item_code': "Item Code SHARE 2021-01-23.xlsx",
        'demographics': "2020 Demographics File Sample 2021-02-17.xlsx",
        'heatmap_color': "Heatmap Colors.xlsx",
        'raw_data_past': "2018 Employee Survey Responses Sample 2021-02-05.xlsx",
        'demographics_past': "2018 Demographics File Sample 2021-02-17.xlsx",
        'benchmark': "External Benchmarks.xlsx",
        'how to use': "How to Use Content 2021-02-28.xlsx",
        'gm_levels': "GM Levels 2021-02-17.xlsx",
        'output_folder': "./output",
        'input_folder': "./input",
        'image': "image.png",
    }

    dfm = DemographicFileMaker(**init_data)
    ## you can change GM_region_human_parentorg here.
    # dfm.setGMParentFlag(0)

    ltm = LTMaker(**init_data)
    ## set GM_region_human_parentorg = 0.
    # ltm.setGMParentFlag(1)

    ## custom text for score summary file.
    init_data.update(
        {
            'custom text': 
"""
* 2018 indicates 2018 Global Employee Survey results (negative indicates less favorable in 2020)
Ext indicates vendor biotechnology & medical device benchmark across companies and geographies
Gilead column compares {}'s scores to Gilead Overall (negative indicates {} less favorable)
"""
        }
    )
    ssm = SSM(**init_data)


    ## read all needed files.
    dfm.readAllFiles()
    ltm.readAllFiles()
    ssm.readAllFiles()

    total_ids = len(dfm.leaders.index) + len(dfm.GMs.index) + len(dfm.site_leads.index)

    victims = []

    for index in tqdm(range(total_ids), desc="total process"):
        if index < len(dfm.leaders.index):
            dfm.setLeader(dfm.leaders.iloc[index, :].values[0])
            ltm.setLeader(ltm.leaders.iloc[index, :].values[0])
            ssm.setLeader(ssm.leaders.iloc[index, :].values[0])

        elif index < len(dfm.leaders.index) + len(dfm.GMs.index):
            row = dfm.GMs.iloc[index - len(dfm.leaders.index), :]
            dfm.setLeader(row["GM ID"], row["GM Org"])

            row = ltm.GMs.iloc[index - len(ltm.leaders.index), :]
            ltm.setLeader(row["GM ID"], row["GM Org"])

            row = ssm.GMs.iloc[index - len(ssm.leaders.index), :]
            ssm.setLeader(row["GM ID"], row["GM Org"])

        else:
            row = dfm.site_leads.iloc[index - len(dfm.leaders.index) - len(dfm.GMs.index), :]
            dfm.setLeader(row["Site Leader ID"], False, row["Site Name"])

            row = ltm.site_leads.iloc[index - len(ltm.leaders.index) - len(ltm.GMs.index), :]
            ltm.setLeader(row["Site Leader ID"], False, row["Site Name"])

            row = ssm.site_leads.iloc[index - len(ssm.leaders.index) - len(ssm.GMs.index), :]
            ssm.setLeader(row["Site Leader ID"], False, row["Site Name"])
    
        ## do main process to calculate report.
        result = dfm.calculateValues()
        if not result:
            ## make report dataframe to output.
            dfm.makeReport()

            ## write output file (Mockup STAR.xlsx)
            book = dfm.getWorkBook()

            ltm.setWorkBook(book)

            ltm.calculateValues()

            ltm.makeReport()

            ltm.writeOutput()

            ssm.calculateValues()

            ssm.makeReport()

            ssm.writeOutput()
        else:
            victims.append([result[0], result[1]])
    
    if len(victims) > 0:
        df = pd.DataFrame(victims, columns=["ID", "Org"])
        df.to_excel(init_data["output_folder"] + "/rest.xlsx", index=False)

    print("complete!")